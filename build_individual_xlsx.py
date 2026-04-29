"""
Build outputs/scenarii_individuale.xlsx — analiza standalone 1 parcare.
Foc principal: SENSIBILITATI — ce parametri muta cel mai mult NPV/IRR.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.comments import Comment
from copy import copy

from tooltips import XLSX_TOOLTIPS

# === Stiluri ===
BLUE = Font(name='Arial', size=10, color='0000FF')
BLACK = Font(name='Arial', size=10, color='000000')
GREEN = Font(name='Arial', size=10, color='008000')
WHITE_BOLD = Font(name='Arial', size=11, bold=True, color='FFFFFF')
SUBHEADER = Font(name='Arial', size=10, bold=True, color='000000')

YELLOW_BG = PatternFill('solid', start_color='FFFF99')
DARK_BLUE_BG = PatternFill('solid', start_color='1F4E78')
LIGHT_BLUE_BG = PatternFill('solid', start_color='D9E1F2')
GRAY_BG = PatternFill('solid', start_color='F2F2F2')
GREEN_BG = PatternFill('solid', start_color='C6EFCE')
RED_BG = PatternFill('solid', start_color='FFC7CE')
ORANGE_BG = PatternFill('solid', start_color='FFD966')

THIN = Side(border_style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')

EUR_FMT = '#,##0 "EUR";(#,##0) "EUR";"-" "EUR"'
RON_FMT = '#,##0.0 "RON"'
PCT_FMT = '0.0%;(0.0%);"-"'
NUM_FMT = '#,##0;(#,##0);"-"'

def style_input(c):
    c.font = BLUE
    c.fill = YELLOW_BG
    c.border = BORDER

def style_formula(c):
    c.font = BLACK
    c.border = BORDER

def style_link(c):
    c.font = GREEN
    c.border = BORDER

def style_header(c):
    c.font = WHITE_BOLD
    c.fill = DARK_BLUE_BG
    c.alignment = CENTER
    c.border = BORDER


# ============================================================
# CALCULE PYTHON pentru pre-popularea sensitivitatilor
# (Excel are formule pentru baseline, dar tornado e pre-calculat in Python pentru claritate)
# ============================================================

EUR_RON = 5.0
ZILE_AN = 365
COTA_IMPOZIT = 0.16
AMORTIZARE_ANI = 5
PROVIZION_PCT = 0.06
ORIZONT_ANI = 5

BUCKETS_RETAIL = [(7.5, 0.25), (22.5, 0.45), (45, 0.22), (90, 0.05), (150, 0.02), (240, 0.01)]


def venit_per_intrare(buckets, gratuitate, tarif_ron_h):
    return sum(p * max(0, m - gratuitate) / 60 * tarif_ron_h for m, p in buckets)


def npv_irr_payback(capex, opex_lunar, venit_an_y1_eur, inflatie, discount, ani=5):
    cf = []
    for an in range(1, ani + 1):
        f = (1 + inflatie) ** (an - 1)
        venit = venit_an_y1_eur * f
        opex = opex_lunar * 12 * f
        ebitda = venit - opex
        provizion = capex * PROVIZION_PCT
        amort = capex / AMORTIZARE_ANI
        prof_imp = ebitda - provizion - amort
        tax = max(0, prof_imp * COTA_IMPOZIT)
        prof_net = prof_imp - tax
        cf_an = prof_net + amort  # cash 100%, fara anuitate
        cf.append(cf_an)
    npv = -capex + sum(c / (1 + discount) ** (i + 1) for i, c in enumerate(cf))
    # IRR
    flows = [-capex] + cf
    irr = None
    if not (all(f >= 0 for f in flows) or all(f <= 0 for f in flows)):
        lo, hi = -0.99, 5.0
        for _ in range(200):
            mid = (lo + hi) / 2
            v = sum(f / (1 + mid) ** i for i, f in enumerate(flows))
            if abs(v) < 1.0:
                irr = mid
                break
            v_lo = sum(f / (1 + lo) ** i for i, f in enumerate(flows))
            if (v_lo > 0 and v > 0) or (v_lo < 0 and v < 0):
                lo = mid
            else:
                hi = mid
        if irr is None:
            irr = mid
    # Payback
    cum = -capex
    payback = None
    for i, c in enumerate(cf, 1):
        cum += c
        if cum >= 0 and payback is None:
            payback = i - (cum - c) / c if c > 0 else i
    return cf, npv, irr, payback


# Baseline RETAIL (parcare mijlocie retail g=120 — cazul marginal)
BASELINE_RETAIL = {
    'CAPEX': 45000,
    'OPEX_lunar': 600,
    'trafic': 1000,
    'tarif_h_RON': 10,
    'gratuitate_min': 120,
    'inflatie': 0.06,
    'discount': 0.12,
}

def venit_an_retail(p):
    v_intr = venit_per_intrare(BUCKETS_RETAIL, p['gratuitate_min'], p['tarif_h_RON'])
    return v_intr * p['trafic'] * ZILE_AN / EUR_RON

# Baseline NON-RETAIL (parcare mijlocie non-retail standalone)
BASELINE_NONRETAIL = {
    'CAPEX': 45000,
    'OPEX_lunar': 600,
    'trafic': 300,
    'tarif_sesiune_RON': 10,
    'colectare': 0.75,
    'inflatie': 0.06,
    'discount': 0.12,
}

def venit_an_nonretail(p):
    return p['tarif_sesiune_RON'] * p['trafic'] * p['colectare'] * ZILE_AN / EUR_RON


# Tornado: pentru fiecare variabila, NPV la valoarea low, baseline, high
def tornado_retail(baseline):
    rows = []
    # CAPEX -25% / +25%
    for low, high, var in [
        (35000, 55000, 'CAPEX (EUR)'),
        (450,   750,   'OPEX lunar (EUR)'),
        (750,   1250,  'Trafic (intrari/zi)'),
        (7.5,   12.5,  'Tarif RON/h'),
    ]:
        for case, val in [('low', low), ('high', high)]:
            p = dict(baseline)
            key_map = {
                'CAPEX (EUR)': 'CAPEX',
                'OPEX lunar (EUR)': 'OPEX_lunar',
                'Trafic (intrari/zi)': 'trafic',
                'Tarif RON/h': 'tarif_h_RON',
            }
            p[key_map[var]] = val
            v_an = venit_an_retail(p)
            _, npv, irr, _ = npv_irr_payback(p['CAPEX'], p['OPEX_lunar'], v_an, p['inflatie'], p['discount'])
        # baseline
        p_base = baseline.copy()
        v_an_base = venit_an_retail(p_base)
        _, npv_base, irr_base, _ = npv_irr_payback(p_base['CAPEX'], p_base['OPEX_lunar'], v_an_base, p_base['inflatie'], p_base['discount'])

        # Compute low/high
        p_low = dict(baseline); p_low[key_map[var]] = low
        p_high = dict(baseline); p_high[key_map[var]] = high
        v_low = venit_an_retail(p_low)
        v_high = venit_an_retail(p_high)
        _, npv_low, irr_low, _ = npv_irr_payback(p_low['CAPEX'], p_low['OPEX_lunar'], v_low, p_low['inflatie'], p_low['discount'])
        _, npv_high, irr_high, _ = npv_irr_payback(p_high['CAPEX'], p_high['OPEX_lunar'], v_high, p_high['inflatie'], p_high['discount'])

        rows.append({
            'var': var,
            'low_val': low,
            'baseline_val': baseline[key_map[var]],
            'high_val': high,
            'npv_low': npv_low,
            'npv_baseline': npv_base,
            'npv_high': npv_high,
            'irr_low': irr_low,
            'irr_high': irr_high,
            'impact': abs(npv_high - npv_low),
        })

    # Gratuitate (discrete)
    for grat in [60, 120, 180]:
        p = dict(baseline)
        p['gratuitate_min'] = grat
        v = venit_an_retail(p)
        _, npv_g, irr_g, _ = npv_irr_payback(p['CAPEX'], p['OPEX_lunar'], v, p['inflatie'], p['discount'])
    # Compute proper low/high for gratuitate
    p_60 = dict(baseline); p_60['gratuitate_min'] = 60
    p_180 = dict(baseline); p_180['gratuitate_min'] = 180
    v_60 = venit_an_retail(p_60); v_180 = venit_an_retail(p_180)
    _, npv_60, irr_60, _ = npv_irr_payback(p_60['CAPEX'], p_60['OPEX_lunar'], v_60, p_60['inflatie'], p_60['discount'])
    _, npv_180, irr_180, _ = npv_irr_payback(p_180['CAPEX'], p_180['OPEX_lunar'], v_180, p_180['inflatie'], p_180['discount'])
    rows.append({
        'var': 'Gratuitate (min)',
        'low_val': 180,
        'baseline_val': 120,
        'high_val': 60,
        'npv_low': npv_180,
        'npv_baseline': npv_base,
        'npv_high': npv_60,
        'irr_low': irr_180,
        'irr_high': irr_60,
        'impact': abs(npv_60 - npv_180),
    })

    # Inflatie
    p_3 = dict(baseline); p_3['inflatie'] = 0.03
    p_9 = dict(baseline); p_9['inflatie'] = 0.09
    v_b = venit_an_retail(p_3)
    _, npv_3, _, _ = npv_irr_payback(p_3['CAPEX'], p_3['OPEX_lunar'], v_b, 0.03, p_3['discount'])
    _, npv_9, _, _ = npv_irr_payback(p_9['CAPEX'], p_9['OPEX_lunar'], v_b, 0.09, p_9['discount'])
    rows.append({
        'var': 'Inflatie (% an)',
        'low_val': 0.03,
        'baseline_val': 0.06,
        'high_val': 0.09,
        'npv_low': npv_3,
        'npv_baseline': npv_base,
        'npv_high': npv_9,
        'irr_low': None,
        'irr_high': None,
        'impact': abs(npv_9 - npv_3),
    })

    # Discount rate
    p_8 = dict(baseline); p_8['discount'] = 0.08
    p_18 = dict(baseline); p_18['discount'] = 0.18
    v_b = venit_an_retail(p_8)
    _, npv_8, _, _ = npv_irr_payback(p_8['CAPEX'], p_8['OPEX_lunar'], v_b, p_8['inflatie'], 0.08)
    _, npv_18, _, _ = npv_irr_payback(p_18['CAPEX'], p_18['OPEX_lunar'], v_b, p_18['inflatie'], 0.18)
    rows.append({
        'var': 'Discount rate (%)',
        'low_val': 0.18,
        'baseline_val': 0.12,
        'high_val': 0.08,
        'npv_low': npv_18,
        'npv_baseline': npv_base,
        'npv_high': npv_8,
        'irr_low': None,
        'irr_high': None,
        'impact': abs(npv_8 - npv_18),
    })

    # Sort by impact (largest first)
    rows.sort(key=lambda r: r['impact'], reverse=True)
    return rows, npv_base, irr_base


def tornado_nonretail(baseline):
    rows = []
    key_map = {
        'CAPEX (EUR)': 'CAPEX',
        'OPEX lunar (EUR)': 'OPEX_lunar',
        'Trafic (intrari/zi)': 'trafic',
        'Tarif RON/sesiune': 'tarif_sesiune_RON',
        'Colectare (%)': 'colectare',
    }

    # baseline
    p_base = baseline.copy()
    v_an_base = venit_an_nonretail(p_base)
    _, npv_base, irr_base, _ = npv_irr_payback(p_base['CAPEX'], p_base['OPEX_lunar'], v_an_base, p_base['inflatie'], p_base['discount'])

    for low, high, var in [
        (35000, 55000, 'CAPEX (EUR)'),
        (450, 750, 'OPEX lunar (EUR)'),
        (200, 400, 'Trafic (intrari/zi)'),
        (7.5, 12.5, 'Tarif RON/sesiune'),
        (0.50, 0.90, 'Colectare (%)'),
    ]:
        p_low = dict(baseline); p_low[key_map[var]] = low
        p_high = dict(baseline); p_high[key_map[var]] = high
        v_low = venit_an_nonretail(p_low)
        v_high = venit_an_nonretail(p_high)
        _, npv_low, irr_low, _ = npv_irr_payback(p_low['CAPEX'], p_low['OPEX_lunar'], v_low, p_low['inflatie'], p_low['discount'])
        _, npv_high, irr_high, _ = npv_irr_payback(p_high['CAPEX'], p_high['OPEX_lunar'], v_high, p_high['inflatie'], p_high['discount'])
        rows.append({
            'var': var,
            'low_val': low,
            'baseline_val': baseline[key_map[var]],
            'high_val': high,
            'npv_low': npv_low,
            'npv_baseline': npv_base,
            'npv_high': npv_high,
            'irr_low': irr_low,
            'irr_high': irr_high,
            'impact': abs(npv_high - npv_low),
        })

    # Inflatie
    p_3 = dict(baseline); p_3['inflatie'] = 0.03
    p_9 = dict(baseline); p_9['inflatie'] = 0.09
    v_b = venit_an_nonretail(p_3)
    _, npv_3, _, _ = npv_irr_payback(p_3['CAPEX'], p_3['OPEX_lunar'], v_b, 0.03, p_3['discount'])
    _, npv_9, _, _ = npv_irr_payback(p_9['CAPEX'], p_9['OPEX_lunar'], v_b, 0.09, p_9['discount'])
    rows.append({
        'var': 'Inflatie (% an)',
        'low_val': 0.03,
        'baseline_val': 0.06,
        'high_val': 0.09,
        'npv_low': npv_3,
        'npv_baseline': npv_base,
        'npv_high': npv_9,
        'irr_low': None,
        'irr_high': None,
        'impact': abs(npv_9 - npv_3),
    })

    # Discount
    p_8 = dict(baseline); p_8['discount'] = 0.08
    p_18 = dict(baseline); p_18['discount'] = 0.18
    v_b = venit_an_nonretail(p_8)
    _, npv_8, _, _ = npv_irr_payback(p_8['CAPEX'], p_8['OPEX_lunar'], v_b, p_8['inflatie'], 0.08)
    _, npv_18, _, _ = npv_irr_payback(p_18['CAPEX'], p_18['OPEX_lunar'], v_b, p_18['inflatie'], 0.18)
    rows.append({
        'var': 'Discount rate (%)',
        'low_val': 0.18,
        'baseline_val': 0.12,
        'high_val': 0.08,
        'npv_low': npv_18,
        'npv_baseline': npv_base,
        'npv_high': npv_8,
        'irr_low': None,
        'irr_high': None,
        'impact': abs(npv_8 - npv_18),
    })

    rows.sort(key=lambda r: r['impact'], reverse=True)
    return rows, npv_base, irr_base


tornado_r, npv_base_r, irr_base_r = tornado_retail(BASELINE_RETAIL)
tornado_nr, npv_base_nr, irr_base_nr = tornado_nonretail(BASELINE_NONRETAIL)


# ============================================================
# BUILD EXCEL
# ============================================================
wb = Workbook()


def section_title(ws, row, col, title, span=4):
    c = ws.cell(row=row, column=col)
    c.value = title
    c.font = WHITE_BOLD
    c.fill = DARK_BLUE_BG
    c.alignment = LEFT
    c.border = BORDER
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)


# ====== SHEET 1: PARAMETRI INDIVIDUALI ======
ws_p = wb.active
ws_p.title = 'Parametri'
ws_p.sheet_properties.tabColor = '4472C4'

ws_p.column_dimensions['A'].width = 36
ws_p.column_dimensions['B'].width = 14
ws_p.column_dimensions['C'].width = 14
ws_p.column_dimensions['D'].width = 50

ws_p['A1'].value = 'PARAMETRI INDIVIDUALI — analiza 1 parcare standalone'
ws_p['A1'].font = Font(name='Arial', size=14, bold=True, color='1F4E78')
ws_p.merge_cells('A1:D1')
ws_p['A2'].value = 'Coloana B = parcare RETAIL (Lidl/Kaufland/etc). Coloana C = parcare NON-RETAIL (standalone, captiv). Modifica orice celula galbena.'
ws_p['A2'].font = Font(name='Arial', size=9, italic=True, color='595959')
ws_p.merge_cells('A2:D2')

NAMED = {}

section_title(ws_p, 4, 1, 'A. PARAMETRI ECONOMICI GLOBALI', 4)
fin = [
    ('curs_eur_ron', 'Curs EUR/RON', 5.0, '0.00'),
    ('cota_impozit', 'Cota impozit profit (SA = 16%)', 0.16, PCT_FMT),
    ('amortizare_ani', 'Amortizare CAPEX (ani)', 5, NUM_FMT),
    ('provizion_pct', 'Provizion CAPEX recurent (% anual)', 0.06, PCT_FMT),
    ('inflatie', 'Inflatie OPEX si tarife (% anual)', 0.06, PCT_FMT),
    ('orizont_ani', 'Orizont analiza (ani)', 5, NUM_FMT),
    ('discount', 'Rata discount (cost capital propriu)', 0.12, PCT_FMT),
]
row = 5
for code, label, val, fmt in fin:
    ws_p.cell(row=row, column=1).value = label
    ws_p.cell(row=row, column=1).border = BORDER
    c = ws_p.cell(row=row, column=2)
    c.value = val
    style_input(c)
    c.number_format = fmt
    NAMED[code] = f"Parametri!$B${row}"
    row += 1

# B. RETAIL
row += 1
section_title(ws_p, row, 1, 'B. PARCARE RETAIL (cu perioada de gratuitate)', 4)
row += 1
retail = [
    ('r_capex',     'CAPEX (EUR)',                        45000, EUR_FMT),
    ('r_opex',      'OPEX lunar (EUR)',                   600,   EUR_FMT),
    ('r_locuri',    'Numar locuri parcare',                150,  NUM_FMT),
    ('r_trafic',    'Trafic intrari/zi',                   1000, NUM_FMT),
    ('r_tarif',     'Tarif (RON/h)',                       10,   '0.0'),
    ('r_grat',      'Perioada gratuitate (min)',           120,  NUM_FMT),
]
RETAIL_ROWS = {}
for code, label, val, fmt in retail:
    ws_p.cell(row=row, column=1).value = label
    ws_p.cell(row=row, column=1).border = BORDER
    c = ws_p.cell(row=row, column=2)
    c.value = val
    style_input(c)
    c.number_format = fmt
    NAMED[code] = f"Parametri!$B${row}"
    RETAIL_ROWS[code] = row
    row += 1

# Distributie durate retail
row += 1
ws_p.cell(row=row, column=1).value = 'Distributie durate stationare retail'
ws_p.cell(row=row, column=1).font = SUBHEADER
ws_p.cell(row=row, column=1).fill = LIGHT_BLUE_BG
ws_p.cell(row=row, column=1).border = BORDER
ws_p.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1
distr_headers = ['Bucket', 'Mid-point (min)', 'Pondere (%)']
for i, h in enumerate(distr_headers):
    c = ws_p.cell(row=row, column=1 + i)
    c.value = h
    c.font = SUBHEADER
    c.fill = GRAY_BG
    c.border = BORDER
row += 1
distr_data = [
    ('0-15 min',     7.5,  0.25),
    ('15-30 min',    22.5, 0.45),
    ('30-60 min',    45,   0.22),
    ('60-120 min',   90,   0.05),
    ('120-180 min',  150,  0.02),
    ('180+ min',     240,  0.01),
]
DISTR_START = row
for bucket, mid, pondere in distr_data:
    ws_p.cell(row=row, column=1).value = bucket
    ws_p.cell(row=row, column=1).border = BORDER
    c = ws_p.cell(row=row, column=2)
    c.value = mid
    style_input(c)
    c.number_format = '0.0'
    c = ws_p.cell(row=row, column=3)
    c.value = pondere
    style_input(c)
    c.number_format = PCT_FMT
    row += 1
DISTR_END = row - 1
# Suma check
ws_p.cell(row=row, column=1).value = 'TOTAL pondere (trebuie 100%)'
ws_p.cell(row=row, column=1).font = SUBHEADER
ws_p.cell(row=row, column=1).fill = GRAY_BG
ws_p.cell(row=row, column=1).border = BORDER
c = ws_p.cell(row=row, column=3)
c.value = f"=SUM(C{DISTR_START}:C{DISTR_END})"
style_formula(c)
c.fill = GRAY_BG
c.number_format = PCT_FMT
row += 1

# C. NON-RETAIL
row += 1
section_title(ws_p, row, 1, 'C. PARCARE NON-RETAIL (fara perioada de gratuitate)', 4)
row += 1
nonretail = [
    ('nr_capex',       'CAPEX (EUR)',                       45000, EUR_FMT),
    ('nr_opex',        'OPEX lunar (EUR)',                  600,   EUR_FMT),
    ('nr_locuri',      'Numar locuri parcare',              100,   NUM_FMT),
    ('nr_trafic',      'Trafic intrari/zi',                 300,   NUM_FMT),
    ('nr_tarif',       'Tarif (RON/sesiune medie 2h)',       10,   '0.0'),
    ('nr_colectare',   'Rata colectare (% platitori)',      0.75,  PCT_FMT),
]
NONRETAIL_ROWS = {}
for code, label, val, fmt in nonretail:
    ws_p.cell(row=row, column=1).value = label
    ws_p.cell(row=row, column=1).border = BORDER
    c = ws_p.cell(row=row, column=2)
    c.value = val
    style_input(c)
    c.number_format = fmt
    NAMED[code] = f"Parametri!$B${row}"
    NONRETAIL_ROWS[code] = row
    row += 1

# Note
row += 1
ws_p.cell(row=row, column=1).value = 'NOTA: tarif RON/sesiune medie. Daca tarif e RON/h, multiplica cu durata medie (2h tipic).'
ws_p.cell(row=row, column=1).font = Font(name='Arial', size=9, italic=True, color='595959')
ws_p.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)


# ====== SHEET 2: CALCUL — P&L 5 ani pentru ambele cazuri ======
ws_c = wb.create_sheet('Calcul')
ws_c.sheet_properties.tabColor = '70AD47'
ws_c.column_dimensions['A'].width = 38
for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
    ws_c.column_dimensions[col_letter].width = 14

ws_c['A1'].value = 'P&L 5 ANI — calcul automat din Parametri'
ws_c['A1'].font = Font(name='Arial', size=14, bold=True, color='1F4E78')
ws_c.merge_cells('A1:G1')


def build_pnl_block(ws, start_row, label_block, capex_ref, opex_ref, venit_an_y1_formula, named):
    """Construieste un bloc P&L pe 5 ani cu inflatie aplicata."""
    # Header bloc
    section_title(ws, start_row, 1, label_block, 7)
    r = start_row + 1
    headers = ['Element', 'An 1', 'An 2', 'An 3', 'An 4', 'An 5', 'Cumulat']
    for i, h in enumerate(headers):
        c = ws.cell(row=r, column=1 + i)
        c.value = h
        style_header(c)
    r += 1
    # Venit
    ws.cell(row=r, column=1).value = 'Venit operational'
    style_formula(ws.cell(row=r, column=1))
    for an in range(1, 6):
        c = ws.cell(row=r, column=1 + an)
        c.value = f"={venit_an_y1_formula}*(1+{named['inflatie']})^({an}-1)"
        style_formula(c)
        c.number_format = EUR_FMT
    c = ws.cell(row=r, column=7)
    c.value = f"=SUM(B{r}:F{r})"
    style_formula(c)
    c.number_format = EUR_FMT
    venit_row = r
    r += 1

    # OPEX direct
    ws.cell(row=r, column=1).value = '(-) OPEX direct'
    style_formula(ws.cell(row=r, column=1))
    for an in range(1, 6):
        c = ws.cell(row=r, column=1 + an)
        c.value = f"=-{opex_ref}*12*(1+{named['inflatie']})^({an}-1)"
        style_formula(c)
        c.number_format = EUR_FMT
    c = ws.cell(row=r, column=7)
    c.value = f"=SUM(B{r}:F{r})"
    style_formula(c)
    c.number_format = EUR_FMT
    opex_row = r
    r += 1

    # EBITDA
    ws.cell(row=r, column=1).value = 'EBITDA'
    ws.cell(row=r, column=1).font = SUBHEADER
    ws.cell(row=r, column=1).fill = LIGHT_BLUE_BG
    ws.cell(row=r, column=1).border = BORDER
    for an in range(1, 6):
        col = 1 + an
        L = get_column_letter(col)
        c = ws.cell(row=r, column=col)
        c.value = f"={L}{venit_row}+{L}{opex_row}"
        style_formula(c)
        c.fill = LIGHT_BLUE_BG
        c.number_format = EUR_FMT
    c = ws.cell(row=r, column=7)
    c.value = f"=SUM(B{r}:F{r})"
    style_formula(c)
    c.fill = LIGHT_BLUE_BG
    c.number_format = EUR_FMT
    ebitda_row = r
    r += 1

    # Provizion
    ws.cell(row=r, column=1).value = '(-) Provizion CAPEX (constant)'
    style_formula(ws.cell(row=r, column=1))
    for an in range(1, 6):
        c = ws.cell(row=r, column=1 + an)
        c.value = f"=-{capex_ref}*{named['provizion_pct']}"
        style_formula(c)
        c.number_format = EUR_FMT
    c = ws.cell(row=r, column=7)
    c.value = f"=SUM(B{r}:F{r})"
    style_formula(c)
    c.number_format = EUR_FMT
    prov_row = r
    r += 1

    # Amortizare
    ws.cell(row=r, column=1).value = '(-) Amortizare (constant)'
    style_formula(ws.cell(row=r, column=1))
    for an in range(1, 6):
        c = ws.cell(row=r, column=1 + an)
        c.value = f"=-{capex_ref}/{named['amortizare_ani']}"
        style_formula(c)
        c.number_format = EUR_FMT
    c = ws.cell(row=r, column=7)
    c.value = f"=SUM(B{r}:F{r})"
    style_formula(c)
    c.number_format = EUR_FMT
    amort_row = r
    r += 1

    # Profit impozabil
    ws.cell(row=r, column=1).value = 'Profit impozabil'
    ws.cell(row=r, column=1).font = SUBHEADER
    ws.cell(row=r, column=1).fill = LIGHT_BLUE_BG
    ws.cell(row=r, column=1).border = BORDER
    for an in range(1, 6):
        col = 1 + an
        L = get_column_letter(col)
        c = ws.cell(row=r, column=col)
        c.value = f"={L}{ebitda_row}+{L}{prov_row}+{L}{amort_row}"
        style_formula(c)
        c.fill = LIGHT_BLUE_BG
        c.number_format = EUR_FMT
    c = ws.cell(row=r, column=7)
    c.value = f"=SUM(B{r}:F{r})"
    style_formula(c)
    c.fill = LIGHT_BLUE_BG
    c.number_format = EUR_FMT
    profit_imp_row = r
    r += 1

    # Impozit
    ws.cell(row=r, column=1).value = '(-) Impozit (16% pe pozitiv)'
    style_formula(ws.cell(row=r, column=1))
    for an in range(1, 6):
        col = 1 + an
        L = get_column_letter(col)
        c = ws.cell(row=r, column=col)
        c.value = f"=-MAX(0,{L}{profit_imp_row}*{named['cota_impozit']})"
        style_formula(c)
        c.number_format = EUR_FMT
    c = ws.cell(row=r, column=7)
    c.value = f"=SUM(B{r}:F{r})"
    style_formula(c)
    c.number_format = EUR_FMT
    tax_row = r
    r += 1

    # Profit net
    ws.cell(row=r, column=1).value = 'Profit net'
    style_formula(ws.cell(row=r, column=1))
    for an in range(1, 6):
        col = 1 + an
        L = get_column_letter(col)
        c = ws.cell(row=r, column=col)
        c.value = f"={L}{profit_imp_row}+{L}{tax_row}"
        style_formula(c)
        c.number_format = EUR_FMT
    c = ws.cell(row=r, column=7)
    c.value = f"=SUM(B{r}:F{r})"
    style_formula(c)
    c.number_format = EUR_FMT
    profit_net_row = r
    r += 1

    # + Amortizare (re-adaugare)
    ws.cell(row=r, column=1).value = '(+) Amortizare (non-cash)'
    style_formula(ws.cell(row=r, column=1))
    for an in range(1, 6):
        col = 1 + an
        L = get_column_letter(col)
        c = ws.cell(row=r, column=col)
        c.value = f"=-{L}{amort_row}"
        style_formula(c)
        c.number_format = EUR_FMT
    r += 1

    # CF Net (cash 100%, fara anuitate)
    ws.cell(row=r, column=1).value = 'CF NET (cash 100%)'
    ws.cell(row=r, column=1).font = WHITE_BOLD
    ws.cell(row=r, column=1).fill = DARK_BLUE_BG
    ws.cell(row=r, column=1).border = BORDER
    cf_net_row = r
    for an in range(1, 6):
        col = 1 + an
        L = get_column_letter(col)
        c = ws.cell(row=r, column=col)
        c.value = f"={L}{profit_net_row}-{L}{amort_row}"
        c.font = WHITE_BOLD
        c.fill = DARK_BLUE_BG
        c.border = BORDER
        c.alignment = RIGHT
        c.number_format = EUR_FMT
    c = ws.cell(row=r, column=7)
    c.value = f"=SUM(B{r}:F{r})"
    c.font = WHITE_BOLD
    c.fill = DARK_BLUE_BG
    c.border = BORDER
    c.alignment = RIGHT
    c.number_format = EUR_FMT
    r += 1

    # CF cumulat
    ws.cell(row=r, column=1).value = 'CF cumulat'
    style_formula(ws.cell(row=r, column=1))
    ws.cell(row=r, column=1).fill = LIGHT_BLUE_BG
    cf_cum_row = r
    ws.cell(row=r, column=2).value = f"=B{cf_net_row}-{capex_ref}"
    style_formula(ws.cell(row=r, column=2))
    ws.cell(row=r, column=2).fill = LIGHT_BLUE_BG
    ws.cell(row=r, column=2).number_format = EUR_FMT
    for an in range(2, 6):
        col = 1 + an
        L = get_column_letter(col)
        L_prev = get_column_letter(col - 1)
        c = ws.cell(row=r, column=col)
        c.value = f"={L_prev}{cf_cum_row}+{L}{cf_net_row}"
        style_formula(c)
        c.fill = LIGHT_BLUE_BG
        c.number_format = EUR_FMT
    r += 2

    # KPI
    ws.cell(row=r, column=1).value = 'INDICATORI'
    ws.cell(row=r, column=1).font = WHITE_BOLD
    ws.cell(row=r, column=1).fill = DARK_BLUE_BG
    ws.cell(row=r, column=1).border = BORDER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1
    # NPV
    ws.cell(row=r, column=1).value = 'NPV @ discount rate'
    style_formula(ws.cell(row=r, column=1))
    npv_formula = (f"=-{capex_ref}+B{cf_net_row}/(1+{named['discount']})^1"
                   f"+C{cf_net_row}/(1+{named['discount']})^2"
                   f"+D{cf_net_row}/(1+{named['discount']})^3"
                   f"+E{cf_net_row}/(1+{named['discount']})^4"
                   f"+F{cf_net_row}/(1+{named['discount']})^5")
    c = ws.cell(row=r, column=2)
    c.value = npv_formula
    style_formula(c)
    c.number_format = EUR_FMT
    c.font = Font(name='Arial', size=10, bold=True, color='1F4E78')
    npv_row = r
    r += 1
    # IRR
    ws.cell(row=r, column=1).value = 'IRR'
    style_formula(ws.cell(row=r, column=1))
    # IRR pe range: -CAPEX, CF1..CF5
    # Construct array: necesita un helper row sau utilizam range cu valori de pe sheet
    # Simplu: construim un rand auxiliar cu fluxurile
    aux_row = r + 5  # row pentru aux
    # IRR formula: trebuie sa avem valorile contiguous
    # Alternativ: folosim RATE sau IRR cu range care includem
    # Folosim helper row jos
    c = ws.cell(row=r, column=2)
    c.value = f"=IFERROR(IRR(B{aux_row}:G{aux_row}),0)"
    style_formula(c)
    c.number_format = PCT_FMT
    c.font = Font(name='Arial', size=10, bold=True, color='1F4E78')
    irr_row = r
    r += 1
    # Payback
    ws.cell(row=r, column=1).value = 'Payback simplu (ani)'
    style_formula(ws.cell(row=r, column=1))
    pb_formula = (f'=IF(B{cf_cum_row}>=0,1,'
                  f'IF(C{cf_cum_row}>=0,2,'
                  f'IF(D{cf_cum_row}>=0,3,'
                  f'IF(E{cf_cum_row}>=0,4,'
                  f'IF(F{cf_cum_row}>=0,5,">5")))))')
    c = ws.cell(row=r, column=2)
    c.value = pb_formula
    style_formula(c)
    c.font = Font(name='Arial', size=10, bold=True, color='1F4E78')
    r += 1
    # CF cumulat 5 ani
    ws.cell(row=r, column=1).value = 'CF Net cumulat 5 ani'
    style_formula(ws.cell(row=r, column=1))
    c = ws.cell(row=r, column=2)
    c.value = f"=G{cf_net_row}"
    style_formula(c)
    c.number_format = EUR_FMT
    c.font = Font(name='Arial', size=10, bold=True, color='1F4E78')
    r += 1
    # Verdict
    ws.cell(row=r, column=1).value = 'Verdict (bazat pe IRR)'
    style_formula(ws.cell(row=r, column=1))
    verdict_formula = (f'=IF(B{irr_row}<0,"NEBUNESC (NPV negativ)",'
                       f'IF(B{irr_row}<0.07,"NEBUNESC (sub depozit)",'
                       f'IF(B{irr_row}<0.12,"MARGINAL",'
                       f'IF(B{irr_row}<0.20,"VIABIL",'
                       f'IF(B{irr_row}<0.35,"FOARTE BUN","EXCELENT")))))')
    c = ws.cell(row=r, column=2)
    c.value = verdict_formula
    c.font = Font(name='Arial', size=11, bold=True, color='006100')
    c.border = BORDER
    r += 2

    # Helper row pentru IRR (ascuns vizual prin format light)
    ws.cell(row=aux_row, column=1).value = 'Helper IRR (an 0..5)'
    ws.cell(row=aux_row, column=1).font = Font(name='Arial', size=8, italic=True, color='999999')
    c = ws.cell(row=aux_row, column=2)
    c.value = f"=-{capex_ref}"
    c.font = Font(name='Arial', size=8, color='999999')
    for an in range(1, 6):
        col = 2 + an
        L_cf = get_column_letter(1 + an)
        c = ws.cell(row=aux_row, column=col)
        c.value = f"={L_cf}{cf_net_row}"
        c.font = Font(name='Arial', size=8, color='999999')

    return r + 1


# Venit retail formula (referinta din Parametri)
# venit = SUMPRODUCT pe distributie
DISTR_RANGE_MID = f"Parametri!$B${DISTR_START}:$B${DISTR_END}"
DISTR_RANGE_POND = f"Parametri!$C${DISTR_START}:$C${DISTR_END}"
venit_intrare_retail = (
    f"SUMPRODUCT(MAX(0,{DISTR_RANGE_MID}-{NAMED['r_grat']})/60*{NAMED['r_tarif']}*{DISTR_RANGE_POND})"
)
# Note: SUMPRODUCT cu MAX nu functioneaza fara array. Folosim alta abordare.
# In Excel, MAX(0,array-scalar) intoarce scalar, nu array. Avem nevoie de IF sau --(array>0)*(array-scalar).
venit_intrare_retail = (
    f"SUMPRODUCT(IF({DISTR_RANGE_MID}>{NAMED['r_grat']},({DISTR_RANGE_MID}-{NAMED['r_grat']})/60*{NAMED['r_tarif']},0),{DISTR_RANGE_POND})"
)
# IF in SUMPRODUCT in Excel modern functioneaza ca array.
# In LibreOffice si Excel 365 ar trebui sa mearga. Tester.

venit_an_y1_retail = f"({venit_intrare_retail})*{NAMED['r_trafic']}*365/{NAMED['curs_eur_ron']}"

next_row = build_pnl_block(
    ws_c, 3,
    'P&L PARCARE RETAIL (cu perioada de gratuitate)',
    NAMED['r_capex'],
    NAMED['r_opex'],
    venit_an_y1_retail,
    NAMED,
)

# Venit non-retail
venit_an_y1_nonretail = f"{NAMED['nr_tarif']}*{NAMED['nr_trafic']}*{NAMED['nr_colectare']}*365/{NAMED['curs_eur_ron']}"

build_pnl_block(
    ws_c, next_row + 2,
    'P&L PARCARE NON-RETAIL (fara gratuitate, plata sesiune)',
    NAMED['nr_capex'],
    NAMED['nr_opex'],
    venit_an_y1_nonretail,
    NAMED,
)


# ====== SHEET 3: TORNADO RETAIL ======
ws_tr = wb.create_sheet('Tornado Retail')
ws_tr.sheet_properties.tabColor = 'C00000'
ws_tr.column_dimensions['A'].width = 28
for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_tr.column_dimensions[col_letter].width = 14

ws_tr['A1'].value = 'TORNADO SENSIBILITATI — PARCARE RETAIL g=120 baseline'
ws_tr['A1'].font = Font(name='Arial', size=14, bold=True, color='1F4E78')
ws_tr.merge_cells('A1:H1')
ws_tr['A2'].value = f'NPV baseline: {npv_base_r:,.0f} EUR. IRR baseline: {irr_base_r*100:.1f}%. Variabilele sortate dupa impact (cea mai mare = sus).'
ws_tr['A2'].font = Font(name='Arial', size=10, italic=True)
ws_tr.merge_cells('A2:H2')

t_headers = ['Variabila', 'Val low', 'Val baseline', 'Val high', 'NPV low', 'NPV baseline', 'NPV high', 'Impact NPV']
for i, h in enumerate(t_headers):
    c = ws_tr.cell(row=4, column=1 + i)
    c.value = h
    style_header(c)

for i, row in enumerate(tornado_r):
    r = 5 + i
    ws_tr.cell(row=r, column=1).value = row['var']
    ws_tr.cell(row=r, column=1).font = SUBHEADER
    ws_tr.cell(row=r, column=1).border = BORDER

    # Format low / baseline / high based on type
    def fmt_val(v, var_name):
        if 'PCT' in var_name.upper() or '%' in var_name:
            return f"{v*100:.1f}%"
        if 'EUR' in var_name:
            return f"{v:,.0f}"
        if 'RON' in var_name:
            return f"{v:.1f}"
        return f"{v:,.0f}" if v >= 100 else f"{v}"

    for col_idx, key in enumerate(['low_val', 'baseline_val', 'high_val'], start=2):
        c = ws_tr.cell(row=r, column=col_idx)
        c.value = fmt_val(row[key], row['var'])
        c.font = BLACK
        c.alignment = CENTER
        c.border = BORDER
    for col_idx, key in enumerate(['npv_low', 'npv_baseline', 'npv_high'], start=5):
        c = ws_tr.cell(row=r, column=col_idx)
        c.value = row[key]
        c.font = BLACK
        c.alignment = RIGHT
        c.border = BORDER
        c.number_format = EUR_FMT
    c = ws_tr.cell(row=r, column=8)
    c.value = row['impact']
    c.font = Font(name='Arial', size=10, bold=True, color='C00000')
    c.alignment = RIGHT
    c.border = BORDER
    c.number_format = EUR_FMT

# Heatmap pe coloana Impact
last_r = 4 + len(tornado_r)
ws_tr.conditional_formatting.add(
    f"H5:H{last_r}",
    ColorScaleRule(start_type='min', start_color='FFFFFF',
                   end_type='max', end_color='C00000')
)


# ====== SHEET 4: TORNADO NON-RETAIL ======
ws_tnr = wb.create_sheet('Tornado Non-Retail')
ws_tnr.sheet_properties.tabColor = 'C00000'
ws_tnr.column_dimensions['A'].width = 28
for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_tnr.column_dimensions[col_letter].width = 14

ws_tnr['A1'].value = 'TORNADO SENSIBILITATI — PARCARE NON-RETAIL standalone baseline'
ws_tnr['A1'].font = Font(name='Arial', size=14, bold=True, color='1F4E78')
ws_tnr.merge_cells('A1:H1')
ws_tnr['A2'].value = f'NPV baseline: {npv_base_nr:,.0f} EUR. IRR baseline: {irr_base_nr*100:.1f}%. Variabilele sortate dupa impact.'
ws_tnr['A2'].font = Font(name='Arial', size=10, italic=True)
ws_tnr.merge_cells('A2:H2')

for i, h in enumerate(t_headers):
    c = ws_tnr.cell(row=4, column=1 + i)
    c.value = h
    style_header(c)

for i, row in enumerate(tornado_nr):
    r = 5 + i
    ws_tnr.cell(row=r, column=1).value = row['var']
    ws_tnr.cell(row=r, column=1).font = SUBHEADER
    ws_tnr.cell(row=r, column=1).border = BORDER
    for col_idx, key in enumerate(['low_val', 'baseline_val', 'high_val'], start=2):
        c = ws_tnr.cell(row=r, column=col_idx)
        v = row[key]
        if isinstance(v, float) and v < 1 and 'Discount' not in row['var'] and 'Inflatie' not in row['var']:
            c.value = f"{v*100:.1f}%"
        elif isinstance(v, float) and ('Discount' in row['var'] or 'Inflatie' in row['var'] or 'Colectare' in row['var']):
            c.value = f"{v*100:.1f}%"
        else:
            c.value = f"{v:,.1f}" if isinstance(v, float) else f"{v}"
        c.font = BLACK
        c.alignment = CENTER
        c.border = BORDER
    for col_idx, key in enumerate(['npv_low', 'npv_baseline', 'npv_high'], start=5):
        c = ws_tnr.cell(row=r, column=col_idx)
        c.value = row[key]
        c.font = BLACK
        c.alignment = RIGHT
        c.border = BORDER
        c.number_format = EUR_FMT
    c = ws_tnr.cell(row=r, column=8)
    c.value = row['impact']
    c.font = Font(name='Arial', size=10, bold=True, color='C00000')
    c.alignment = RIGHT
    c.border = BORDER
    c.number_format = EUR_FMT

last_r = 4 + len(tornado_nr)
ws_tnr.conditional_formatting.add(
    f"H5:H{last_r}",
    ColorScaleRule(start_type='min', start_color='FFFFFF',
                   end_type='max', end_color='C00000')
)


# ====== SHEET 5: PRAGURI ======
ws_pr = wb.create_sheet('Praguri')
ws_pr.sheet_properties.tabColor = 'FFC000'
ws_pr.column_dimensions['A'].width = 35
for col_letter in ['B', 'C', 'D']:
    ws_pr.column_dimensions[col_letter].width = 18

ws_pr['A1'].value = 'PRAGURI MINIMA PENTRU VIABILITATE'
ws_pr['A1'].font = Font(name='Arial', size=14, bold=True, color='1F4E78')
ws_pr.merge_cells('A1:D1')
ws_pr['A2'].value = 'La ce valori minime ale fiecarui parametru investitia devine viabila (IRR >= 15%)?'
ws_pr['A2'].font = Font(name='Arial', size=10, italic=True)
ws_pr.merge_cells('A2:D2')

ws_pr['A4'].value = 'Categorie / parametru'
style_header(ws_pr['A4'])
ws_pr['B4'].value = 'CAPEX 25k (mica)'
style_header(ws_pr['B4'])
ws_pr['C4'].value = 'CAPEX 45k (medie)'
style_header(ws_pr['C4'])
ws_pr['D4'].value = 'Note'
style_header(ws_pr['D4'])


def break_even_traffic(capex, opex_an, venit_per_intrare_eur, prag_irr=0.15, ani=5):
    if venit_per_intrare_eur <= 0:
        return None
    annuity_factor = prag_irr / (1 - (1 + prag_irr) ** -ani)
    cf_target = capex * annuity_factor
    amort = capex / AMORTIZARE_ANI
    provizion = capex * PROVIZION_PCT
    venit_target_an = (cf_target - amort) / (1 - COTA_IMPOZIT) + opex_an + provizion
    return venit_target_an / (venit_per_intrare_eur * ZILE_AN)


pragurile_data = []
# Retail praguri
for grat, label in [(120, 'Retail g=120 — trafic minim/zi'), (60, 'Retail g=60 — trafic minim/zi'), (30, 'Retail g=30 — trafic minim/zi')]:
    v_intr_eur = venit_per_intrare(BUCKETS_RETAIL, grat, 10) / EUR_RON
    if v_intr_eur > 0:
        be_25 = break_even_traffic(25000, 450 * 12, v_intr_eur)
        be_45 = break_even_traffic(45000, 600 * 12, v_intr_eur)
        pragurile_data.append((label, f"{be_25:.0f} intrari/zi", f"{be_45:.0f} intrari/zi", f"Tarif 10 RON/h, distributie standard"))
    else:
        pragurile_data.append((label, "imposibil", "imposibil", "Niciun bucket nu plateste"))

# Tarif retail prag
def break_even_tarif(capex, opex_an, gratuitate, trafic, prag_irr=0.15, ani=5):
    annuity_factor = prag_irr / (1 - (1 + prag_irr) ** -ani)
    cf_target = capex * annuity_factor
    amort = capex / AMORTIZARE_ANI
    provizion = capex * PROVIZION_PCT
    venit_target_an = (cf_target - amort) / (1 - COTA_IMPOZIT) + opex_an + provizion
    venit_target_intrare_ron = venit_target_an * EUR_RON / (trafic * ZILE_AN)
    # venit_intrare = SUM(p_i × max(0, m_i - grat)/60 × tarif)
    # So tarif_min = venit_intrare / SUM(p_i × max(0, m_i - grat)/60)
    coef = sum(p * max(0, m - gratuitate) / 60 for m, p in BUCKETS_RETAIL)
    if coef <= 0:
        return None
    return venit_target_intrare_ron / coef


for grat, label in [(120, 'Retail g=120 — tarif minim RON/h'), (60, 'Retail g=60 — tarif minim RON/h')]:
    t25 = break_even_tarif(25000, 450 * 12, grat, 500)
    t45 = break_even_tarif(45000, 600 * 12, grat, 1000)
    pragurile_data.append((label, f"{t25:.1f} RON/h" if t25 else "imposibil", f"{t45:.1f} RON/h" if t45 else "imposibil", f"Trafic 500 (mic) / 1000 (medie) intrari/zi"))

# Non-retail praguri
def break_even_colectare(capex, opex_an, tarif_eur, trafic, prag_irr=0.15, ani=5):
    annuity_factor = prag_irr / (1 - (1 + prag_irr) ** -ani)
    cf_target = capex * annuity_factor
    amort = capex / AMORTIZARE_ANI
    provizion = capex * PROVIZION_PCT
    venit_target_an = (cf_target - amort) / (1 - COTA_IMPOZIT) + opex_an + provizion
    return venit_target_an / (tarif_eur * trafic * ZILE_AN)


def break_even_trafic_nonretail(capex, opex_an, tarif_eur, colectare, prag_irr=0.15, ani=5):
    annuity_factor = prag_irr / (1 - (1 + prag_irr) ** -ani)
    cf_target = capex * annuity_factor
    amort = capex / AMORTIZARE_ANI
    provizion = capex * PROVIZION_PCT
    venit_target_an = (cf_target - amort) / (1 - COTA_IMPOZIT) + opex_an + provizion
    return venit_target_an / (tarif_eur * colectare * ZILE_AN)


tarif_eur = 10 / EUR_RON  # 2 EUR/sesiune
for col, label in [(0.50, 'Non-retail col=50% — trafic minim/zi'), (0.75, 'Non-retail col=75% — trafic minim/zi')]:
    t25 = break_even_trafic_nonretail(25000, 450 * 12, tarif_eur, col)
    t45 = break_even_trafic_nonretail(45000, 600 * 12, tarif_eur, col)
    pragurile_data.append((label, f"{t25:.0f} intrari/zi", f"{t45:.0f} intrari/zi", "Tarif 10 RON/sesiune"))

for trafic, label in [(150, 'Non-retail trafic=150 — colectare minima'), (300, 'Non-retail trafic=300 — colectare minima')]:
    c25 = break_even_colectare(25000, 450 * 12, tarif_eur, trafic)
    c45 = break_even_colectare(45000, 600 * 12, tarif_eur, trafic)
    pragurile_data.append((label, f"{c25*100:.1f}%" if c25 < 1 else "imposibil", f"{c45*100:.1f}%" if c45 < 1 else "imposibil", "Tarif 10 RON/sesiune"))

# Tarif non-retail minim
def break_even_tarif_nonretail(capex, opex_an, trafic, colectare, prag_irr=0.15, ani=5):
    annuity_factor = prag_irr / (1 - (1 + prag_irr) ** -ani)
    cf_target = capex * annuity_factor
    amort = capex / AMORTIZARE_ANI
    provizion = capex * PROVIZION_PCT
    venit_target_an = (cf_target - amort) / (1 - COTA_IMPOZIT) + opex_an + provizion
    return venit_target_an * EUR_RON / (trafic * colectare * ZILE_AN)


for trafic, col, label in [(200, 0.65, 'Tarif minim non-retail (200/zi, 65%)'), (300, 0.75, 'Tarif minim non-retail (300/zi, 75%)')]:
    t25 = break_even_tarif_nonretail(25000, 450 * 12, trafic, col)
    t45 = break_even_tarif_nonretail(45000, 600 * 12, trafic, col)
    pragurile_data.append((label, f"{t25:.1f} RON/sesiune", f"{t45:.1f} RON/sesiune", "Pragul de viabilitate IRR>=15%"))


for i, (label, b, c, note) in enumerate(pragurile_data):
    r = 5 + i
    ws_pr.cell(row=r, column=1).value = label
    ws_pr.cell(row=r, column=1).font = SUBHEADER
    ws_pr.cell(row=r, column=1).border = BORDER
    ws_pr.cell(row=r, column=2).value = b
    ws_pr.cell(row=r, column=2).font = BLACK
    ws_pr.cell(row=r, column=2).alignment = CENTER
    ws_pr.cell(row=r, column=2).border = BORDER
    ws_pr.cell(row=r, column=3).value = c
    ws_pr.cell(row=r, column=3).font = BLACK
    ws_pr.cell(row=r, column=3).alignment = CENTER
    ws_pr.cell(row=r, column=3).border = BORDER
    ws_pr.cell(row=r, column=4).value = note
    ws_pr.cell(row=r, column=4).font = Font(name='Arial', size=9, italic=True)
    ws_pr.cell(row=r, column=4).border = BORDER


# ============================================================
# TOOLTIPS — atasam Comment openpyxl pe celulele cu label-uri cunoscute.
# In Excel/LibreOffice apar pe hover (triunghi rosu in colt celula).
# ============================================================
def attach_all_tooltips(workbook):
    matched = 0
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                tip = XLSX_TOOLTIPS.get(cell.value.strip())
                if not tip or cell.comment is not None:
                    continue
                comment = Comment(tip, "Analiza")
                comment.width = 360
                comment.height = 160
                cell.comment = comment
                matched += 1
    return matched


tooltip_count = attach_all_tooltips(wb)

# Save
output = '/Users/home-felix/Total Hub/Analiza fezabilitate parcari de inchiriat/outputs/scenarii_individuale.xlsx'
wb.save(output)
print(f"Tooltip-uri atasate: {tooltip_count}")
print(f"Salvat: {output}")
print(f"Tornado retail rows: {len(tornado_r)}")
print(f"Tornado non-retail rows: {len(tornado_nr)}")
print(f"NPV baseline retail: {npv_base_r:,.0f} EUR, IRR {irr_base_r*100:.2f}%")
print(f"NPV baseline non-retail: {npv_base_nr:,.0f} EUR, IRR {irr_base_nr*100:.2f}%")
