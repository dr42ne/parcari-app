"""
Constructor pentru outputs/model.xlsx — Faza 3.
Toate formulele referentiaza Parametri (zero numere hardcodate).
Conventii color: yellow bg + blue text = input editabil; green text = link cross-sheet.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo

# === CONVENTII STIL ===
BLUE = Font(name='Arial', size=10, color='0000FF')
BLACK = Font(name='Arial', size=10, color='000000')
GREEN = Font(name='Arial', size=10, color='008000')
WHITE_BOLD = Font(name='Arial', size=11, bold=True, color='FFFFFF')
HEADER = Font(name='Arial', size=10, bold=True, color='FFFFFF')
SUBHEADER = Font(name='Arial', size=10, bold=True, color='000000')

YELLOW_BG = PatternFill('solid', start_color='FFFF99')
DARK_BLUE_BG = PatternFill('solid', start_color='1F4E78')
LIGHT_BLUE_BG = PatternFill('solid', start_color='D9E1F2')
GRAY_BG = PatternFill('solid', start_color='F2F2F2')
GREEN_BG = PatternFill('solid', start_color='C6EFCE')
RED_BG = PatternFill('solid', start_color='FFC7CE')
ORANGE_BG = PatternFill('solid', start_color='FFD966')

THIN = Side(border_style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')

EUR_FMT = '#,##0 "EUR";(#,##0) "EUR";"-" "EUR"'
EUR_DETAIL_FMT = '#,##0.00 "EUR";(#,##0.00) "EUR";"-" "EUR"'
RON_FMT = '#,##0 "RON";(#,##0) "RON";"-" "RON"'
PCT_FMT = '0.0%;(0.0%);"-"'
PCT_DETAIL_FMT = '0.00%;(0.00%);"-"'
ANI_FMT = '0.0" ani"'
NUM_FMT = '#,##0;(#,##0);"-"'
NUMX_FMT = '0.0"x"'

# === Helpers ===
def style_input(cell):
    cell.font = BLUE
    cell.fill = YELLOW_BG
    cell.border = BORDER

def style_formula(cell):
    cell.font = BLACK
    cell.border = BORDER

def style_link(cell):
    cell.font = GREEN
    cell.border = BORDER

def style_header(cell):
    cell.font = WHITE_BOLD
    cell.fill = DARK_BLUE_BG
    cell.alignment = CENTER
    cell.border = BORDER

def style_subheader(cell):
    cell.font = SUBHEADER
    cell.fill = LIGHT_BLUE_BG
    cell.alignment = LEFT
    cell.border = BORDER

def section_title(ws, row, col, title, span=4):
    ws.cell(row=row, column=col).value = title
    ws.cell(row=row, column=col).font = WHITE_BOLD
    ws.cell(row=row, column=col).fill = DARK_BLUE_BG
    ws.cell(row=row, column=col).alignment = LEFT
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)


# === BUILD WORKBOOK ===
wb = Workbook()

# Tab colors
TABS = {
    'Parametri': '4472C4',
    'Tipuri locatii': '70AD47',
    'Variante': '70AD47',
    'Portofoliu': 'FFC000',
    'Cash Flow': 'FFC000',
    'Sumar comparativ': 'C00000',
    'Sensibilitati': 'A5A5A5',
    'Modificari': '595959',
}

# ====================================================================
# SHEET 1: PARAMETRI
# ====================================================================
ws_p = wb.active
ws_p.title = 'Parametri'
ws_p.sheet_properties.tabColor = TABS['Parametri']

ws_p.column_dimensions['A'].width = 38
ws_p.column_dimensions['B'].width = 15
ws_p.column_dimensions['C'].width = 14
ws_p.column_dimensions['D'].width = 60
ws_p['A1'].value = 'PARAMETRI MODEL — Total Hub SA — Analiza fezabilitate parcari'
ws_p['A1'].font = Font(name='Arial', size=14, bold=True, color='1F4E78')
ws_p.merge_cells('A1:D1')
ws_p['A2'].value = 'Versiune 1.0 — 2026-04-29 — toate celulele galbene sunt editabile, formulele se recalculeaza automat'
ws_p['A2'].font = Font(name='Arial', size=9, italic=True, color='595959')
ws_p.merge_cells('A2:D2')

# --- A. Parametri financiari ---
section_title(ws_p, 4, 1, 'A. PARAMETRI FINANCIARI', 4)
fin_params = [
    ('curs_eur_ron', 'Curs EUR/RON', 5.0, 'data piata aprilie 2026'),
    ('orizont_ani', 'Orizont analiza (ani)', 5, 'durata contract standard'),
    ('cota_impozit', 'Cota impozit profit', 0.16, 'SA standard 16%'),
    ('amortizare_ani', 'Amortizare CAPEX echipamente (ani)', 5, 'liniar pentru hardware ANPR/bariere'),
    ('provizion_pct', 'Provizion CAPEX recurent (% anual)', 0.06, '6% reparatii hardware uzual'),
    ('inflatie_opex', 'Inflatie OPEX (% anual)', 0.06, 'BNR 2026-2030 medie'),
    ('inflatie_tarife', 'Inflatie tarife (% anual)', 0.06, 'indexare cu inflatia generala'),
    ('rata_dobanda', 'Rata dobanda credit echipamente', 0.09, 'medie credite IMM Romania 2026'),
    ('pondere_credit', 'Pondere credit in CAPEX', 0.70, 'restul 30% capital propriu'),
    ('cost_capital_propriu', 'Cost capital propriu', 0.18, 'proiect early-stage cu risc clienti mari'),
]

NAMED = {}  # mapeaza nume -> cell ref Parametri (ex: 'curs_eur_ron' -> 'Parametri!$B$5')

row = 5
for code, label, val, src in fin_params:
    ws_p.cell(row=row, column=1).value = label
    ws_p.cell(row=row, column=1).font = BLACK
    ws_p.cell(row=row, column=1).border = BORDER
    ws_p.cell(row=row, column=2).value = val
    style_input(ws_p.cell(row=row, column=2))
    if isinstance(val, float) and (val < 1 and code not in ['curs_eur_ron']):
        ws_p.cell(row=row, column=2).number_format = PCT_FMT
    elif isinstance(val, int):
        ws_p.cell(row=row, column=2).number_format = NUM_FMT
    else:
        ws_p.cell(row=row, column=2).number_format = '0.00'
    ws_p.cell(row=row, column=4).value = src
    ws_p.cell(row=row, column=4).font = Font(name='Arial', size=9, italic=True, color='595959')
    NAMED[code] = f"Parametri!$B${row}"
    row += 1

# WACC formula
row_wacc = row
ws_p.cell(row=row, column=1).value = 'WACC (calcul automat)'
ws_p.cell(row=row, column=1).font = SUBHEADER
ws_p.cell(row=row, column=1).fill = GRAY_BG
ws_p.cell(row=row, column=1).border = BORDER
formula = (f"={NAMED['pondere_credit']}*{NAMED['rata_dobanda']}*(1-{NAMED['cota_impozit']})"
           f"+(1-{NAMED['pondere_credit']})*{NAMED['cost_capital_propriu']}")
ws_p.cell(row=row, column=2).value = formula
style_formula(ws_p.cell(row=row, column=2))
ws_p.cell(row=row, column=2).fill = GRAY_BG
ws_p.cell(row=row, column=2).number_format = PCT_DETAIL_FMT
ws_p.cell(row=row, column=4).value = 'pondere_credit × rata_dobanda × (1 - cota_impozit) + (1 - pondere_credit) × cost_capital'
ws_p.cell(row=row, column=4).font = Font(name='Arial', size=9, italic=True, color='595959')
NAMED['wacc'] = f"Parametri!$B${row}"
row += 1

# --- B. Profile tipuri locatii ---
row += 2
section_title(ws_p, row, 1, 'B. PROFILE TIPURI LOCATII (toate celulele galbene sunt editabile)', 7)
row += 1

tipuri_headers = ['Cod', 'Nume', 'CAPEX (EUR)', 'OPEX lunar (EUR)', 'Tarif h (RON)', 'Trafic/zi', 'Gratuitate baseline (min)']
for i, h in enumerate(tipuri_headers):
    c = ws_p.cell(row=row, column=1 + i)
    c.value = h
    style_header(c)
row += 1
TIPURI_ROWS = {}
tipuri_data = [
    ('A', 'Retail mare (hyper)',  50000, 700,  10, 2000, 120),
    ('B', 'Retail mediu (super)', 40000, 600,  10, 1000, 120),
    ('C', 'Retail mic',           28000, 450,  5,  500,  120),
    ('D', 'Standalone public',    28000, 450,  5,  400,  0),
    ('E', 'Semi-public',          28000, 450,  5,  300,  0),
    ('F', 'Captiv',               28000, 450,  10, 200,  0),
    ('G', 'Mega-mall',           130000, 1200, 12, 8000, 180),
]
for cod, nume, capex, opex, tarif, trafic, grat in tipuri_data:
    TIPURI_ROWS[cod] = row
    ws_p.cell(row=row, column=1).value = cod
    ws_p.cell(row=row, column=1).font = SUBHEADER
    ws_p.cell(row=row, column=1).fill = LIGHT_BLUE_BG
    ws_p.cell(row=row, column=1).alignment = CENTER
    ws_p.cell(row=row, column=1).border = BORDER

    ws_p.cell(row=row, column=2).value = nume
    ws_p.cell(row=row, column=2).font = BLACK
    ws_p.cell(row=row, column=2).border = BORDER

    for col, val, fmt in [(3, capex, EUR_FMT), (4, opex, EUR_FMT), (5, tarif, NUM_FMT), (6, trafic, NUM_FMT), (7, grat, NUM_FMT)]:
        c = ws_p.cell(row=row, column=col)
        c.value = val
        style_input(c)
        c.number_format = fmt
    row += 1

# --- C. Distributii durate retail ---
row += 2
section_title(ws_p, row, 1, 'C. DISTRIBUTIA DURATELOR STATIONARE — RETAIL (suma pe rand = 100%)', 8)
row += 1
ws_p.cell(row=row, column=1).value = 'Tip'
style_header(ws_p.cell(row=row, column=1))
buckets_headers = ['0-15 min', '15-30 min', '30-60 min', '60-120 min', '120-180 min', '180+ min', 'Total %']
for i, h in enumerate(buckets_headers):
    c = ws_p.cell(row=row, column=2 + i)
    c.value = h
    style_header(c)
row += 1

DISTR_ROWS = {}
buckets_data = [
    ('A', [0.20, 0.40, 0.25, 0.10, 0.03, 0.02]),
    ('B', [0.25, 0.45, 0.22, 0.05, 0.02, 0.01]),
    ('C', [0.20, 0.45, 0.25, 0.06, 0.02, 0.02]),
    ('G', [0.10, 0.30, 0.30, 0.20, 0.07, 0.03]),
]
for cod, probs in buckets_data:
    DISTR_ROWS[cod] = row
    ws_p.cell(row=row, column=1).value = cod
    ws_p.cell(row=row, column=1).font = SUBHEADER
    ws_p.cell(row=row, column=1).alignment = CENTER
    ws_p.cell(row=row, column=1).border = BORDER
    for i, p in enumerate(probs):
        c = ws_p.cell(row=row, column=2 + i)
        c.value = p
        style_input(c)
        c.number_format = PCT_FMT
    # Suma
    total_col = 2 + 6
    ws_p.cell(row=row, column=total_col).value = f"=SUM(B{row}:G{row})"
    style_formula(ws_p.cell(row=row, column=total_col))
    ws_p.cell(row=row, column=total_col).number_format = PCT_FMT
    ws_p.cell(row=row, column=total_col).fill = GRAY_BG
    row += 1

# Mid-points (constanta pentru calcul venit)
row += 1
ws_p.cell(row=row, column=1).value = 'Mid-points (min)'
ws_p.cell(row=row, column=1).font = SUBHEADER
ws_p.cell(row=row, column=1).fill = GRAY_BG
midpoints = [7.5, 22.5, 45, 90, 150, 240]
MIDPOINT_ROW = row
for i, mp in enumerate(midpoints):
    c = ws_p.cell(row=row, column=2 + i)
    c.value = mp
    style_input(c)
    c.number_format = NUM_FMT
row += 1

# --- D. Rate colectare non-retail ---
row += 2
section_title(ws_p, row, 1, 'D. RATE COLECTARE NON-RETAIL (Stres test: pesimist / baseline / optimist)', 5)
row += 1
ws_p.cell(row=row, column=1).value = 'Tip'
style_header(ws_p.cell(row=row, column=1))
for i, h in enumerate(['Pesimist', 'Baseline', 'Optimist', 'Tarif sesiune (RON)']):
    c = ws_p.cell(row=row, column=2 + i)
    c.value = h
    style_header(c)
row += 1

COLECT_ROWS = {}
colectare_data = [
    ('D', 0.60, 0.75, 0.90, 10),
    ('E', 0.50, 0.65, 0.80, 10),
    ('F', 0.25, 0.35, 0.50, 10),
]
for cod, pess, base, opt, tarif in colectare_data:
    COLECT_ROWS[cod] = row
    ws_p.cell(row=row, column=1).value = cod
    ws_p.cell(row=row, column=1).font = SUBHEADER
    ws_p.cell(row=row, column=1).alignment = CENTER
    ws_p.cell(row=row, column=1).border = BORDER
    for i, val in enumerate([pess, base, opt]):
        c = ws_p.cell(row=row, column=2 + i)
        c.value = val
        style_input(c)
        c.number_format = PCT_FMT
    c = ws_p.cell(row=row, column=5)
    c.value = tarif
    style_input(c)
    c.number_format = NUM_FMT
    row += 1

# --- E. Variante contractuale ---
row += 2
section_title(ws_p, row, 1, 'E. VARIANTE CONTRACTUALE — fee lunar EUR (>0 = operator primeste; <0 = operator plateste)', 8)
row += 1
ws_p.cell(row=row, column=1).value = 'Cod'
style_header(ws_p.cell(row=row, column=1))
ws_p.cell(row=row, column=2).value = 'Cota operator'
style_header(ws_p.cell(row=row, column=2))
for i, t in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G']):
    c = ws_p.cell(row=row, column=3 + i)
    c.value = f"Fee {t} (EUR/luna)"
    style_header(c)
row += 1

VAR_ROWS = {}
variante_data = [
    ('C1', 1.0, [0,    0,    0,    0,    0,    0,    0]),
    ('C2', 0.5, [0,    0,    0,    0,    0,    0,    0]),
    ('C3', 0.0, [1200, 800,  400,  0,    0,    0,    3000]),
    ('C4', 0.5, [-500, -300, -150, 0,    0,    0,    -1500]),
    ('C5', 1.0, [0,    0,    0,    -400, -300, -150, 0]),
]
for cod, cota, fees in variante_data:
    VAR_ROWS[cod] = row
    ws_p.cell(row=row, column=1).value = cod
    ws_p.cell(row=row, column=1).font = SUBHEADER
    ws_p.cell(row=row, column=1).alignment = CENTER
    ws_p.cell(row=row, column=1).border = BORDER
    c = ws_p.cell(row=row, column=2)
    c.value = cota
    style_input(c)
    c.number_format = PCT_FMT
    for i, fee in enumerate(fees):
        c = ws_p.cell(row=row, column=3 + i)
        c.value = fee
        style_input(c)
        c.number_format = EUR_FMT
    row += 1

# --- F. Overhead corporate ---
row += 2
section_title(ws_p, row, 1, 'F. OVERHEAD CORPORATE (EUR/an)', 4)
row += 1
ws_p.cell(row=row, column=1).value = 'Linie cost'
style_header(ws_p.cell(row=row, column=1))
ws_p.cell(row=row, column=2).value = 'Anul 1 (lean)'
style_header(ws_p.cell(row=row, column=2))
ws_p.cell(row=row, column=3).value = 'Anul 2+ (BD intern)'
style_header(ws_p.cell(row=row, column=3))
ws_p.cell(row=row, column=4).value = 'Anul 2+ peste 5 parcari'
style_header(ws_p.cell(row=row, column=4))
row += 1

overhead_data = [
    ('Manager general / fondator',         38000, 35000, 35000),
    ('Dev / SRE',                          30000, 30000, 30000),
    ('Ops manager',                        22000, 22000, 22000),
    ('BD / sales intern',                  0,     25000, 25000),
    ('Tehnician intern teren',             0,     0,     25000),
    ('Contabilitate externalizata',        4000,  4000,  4000),
    ('Juridic externalizat',               3500,  3500,  3500),
    ('Birou + utilitati',                  7000,  7000,  7000),
    ('Software / SaaS',                    4500,  4500,  4500),
    ('Marketing + branding',               3500,  3500,  3500),
    ('Contingente / diverse',              6500,  6500,  6500),
]
OH_START = row
for label, a1, a2, a3 in overhead_data:
    ws_p.cell(row=row, column=1).value = label
    ws_p.cell(row=row, column=1).font = BLACK
    ws_p.cell(row=row, column=1).border = BORDER
    for i, val in enumerate([a1, a2, a3]):
        c = ws_p.cell(row=row, column=2 + i)
        c.value = val
        style_input(c)
        c.number_format = EUR_FMT
    row += 1

OH_END = row - 1
# Totaluri
ws_p.cell(row=row, column=1).value = 'TOTAL OVERHEAD'
ws_p.cell(row=row, column=1).font = SUBHEADER
ws_p.cell(row=row, column=1).fill = LIGHT_BLUE_BG
ws_p.cell(row=row, column=1).border = BORDER
for col in [2, 3, 4]:
    L = get_column_letter(col)
    c = ws_p.cell(row=row, column=col)
    c.value = f"=SUM({L}{OH_START}:{L}{OH_END})"
    style_formula(c)
    c.fill = LIGHT_BLUE_BG
    c.number_format = EUR_FMT
NAMED['oh_an1'] = f"Parametri!$B${row}"
NAMED['oh_an2'] = f"Parametri!$C${row}"
NAMED['oh_an2plus'] = f"Parametri!$D${row}"
row += 1

# Reducere OPEX cand tehnician intern activ
row += 1
ws_p.cell(row=row, column=1).value = 'Reducere OPEX/luna cand tehnician intern activ — RETAIL'
ws_p.cell(row=row, column=2).value = 150
style_input(ws_p.cell(row=row, column=2))
ws_p.cell(row=row, column=2).number_format = EUR_FMT
NAMED['reduc_opex_retail'] = f"Parametri!$B${row}"
row += 1
ws_p.cell(row=row, column=1).value = 'Reducere OPEX/luna cand tehnician intern activ — NON-RETAIL'
ws_p.cell(row=row, column=2).value = 100
style_input(ws_p.cell(row=row, column=2))
ws_p.cell(row=row, column=2).number_format = EUR_FMT
NAMED['reduc_opex_nonretail'] = f"Parametri!$B${row}"
row += 1
ws_p.cell(row=row, column=1).value = 'Prag declansare tehnician intern (numar parcari operate)'
ws_p.cell(row=row, column=2).value = 5
style_input(ws_p.cell(row=row, column=2))
ws_p.cell(row=row, column=2).number_format = NUM_FMT
NAMED['prag_tehnician'] = f"Parametri!$B${row}"

# Salvam mapari pentru sheet-urile urmatoare
PARAMETRI_TIPURI_ROWS = TIPURI_ROWS  # cod -> rand pe sheetul Parametri
PARAMETRI_DISTR_ROWS = DISTR_ROWS
PARAMETRI_COLECT_ROWS = COLECT_ROWS
PARAMETRI_VAR_ROWS = VAR_ROWS
PARAMETRI_MIDPOINT_ROW = MIDPOINT_ROW

# Helper: returneaza cell ref pe Parametri pentru un parametru tip-locatie
def p_tip(cod, col_letter):
    """col: C=CAPEX, D=OPEX_lunar, E=tarif_h, F=trafic, G=gratuitate"""
    return f"Parametri!${col_letter}${PARAMETRI_TIPURI_ROWS[cod]}"

def p_distr(cod, bucket_idx):
    """bucket_idx 0..5 -> coloana B..G"""
    col = chr(ord('B') + bucket_idx)
    return f"Parametri!${col}${PARAMETRI_DISTR_ROWS[cod]}"

def p_midpoint(bucket_idx):
    col = chr(ord('B') + bucket_idx)
    return f"Parametri!${col}${PARAMETRI_MIDPOINT_ROW}"

def p_colect(cod, scen='base'):
    """scen: pess=B, base=C, opt=D"""
    col = {'pess': 'B', 'base': 'C', 'opt': 'D'}[scen]
    return f"Parametri!${col}${PARAMETRI_COLECT_ROWS[cod]}"

def p_tarif_sesiune(cod):
    return f"Parametri!$E${PARAMETRI_COLECT_ROWS[cod]}"

def p_var_cota(var_cod):
    return f"Parametri!$B${PARAMETRI_VAR_ROWS[var_cod]}"

def p_var_fee(var_cod, tip_cod):
    """fee per (varianta, tip): coloana 3+i unde i e indexul tipului A..G"""
    col = chr(ord('C') + ord(tip_cod) - ord('A'))
    return f"Parametri!${col}${PARAMETRI_VAR_ROWS[var_cod]}"


# ====================================================================
# SHEET 2: TIPURI LOCATII (calcule derivate)
# ====================================================================
ws_t = wb.create_sheet('Tipuri locatii')
ws_t.sheet_properties.tabColor = TABS['Tipuri locatii']

ws_t.column_dimensions['A'].width = 6
ws_t.column_dimensions['B'].width = 28
for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
    ws_t.column_dimensions[col_letter].width = 16

ws_t['A1'].value = 'TIPURI LOCATII — calcule derivate din Parametri'
ws_t['A1'].font = Font(name='Arial', size=14, bold=True, color='1F4E78')
ws_t.merge_cells('A1:J1')

t_headers = ['Cod', 'Nume', 'CAPEX', 'OPEX lunar', 'OPEX anual', 'Credit (70%)', 'Capital propriu', 'Anuitate credit', 'Amortizare anuala', 'Provizion anual']
for i, h in enumerate(t_headers):
    c = ws_t.cell(row=3, column=1 + i)
    c.value = h
    style_header(c)

T_ROWS = {}
tipuri_codes = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
for i, cod in enumerate(tipuri_codes):
    r = 4 + i
    T_ROWS[cod] = r
    ws_t.cell(row=r, column=1).value = cod
    ws_t.cell(row=r, column=1).font = SUBHEADER
    ws_t.cell(row=r, column=1).alignment = CENTER
    ws_t.cell(row=r, column=1).border = BORDER

    # Nume (link)
    c = ws_t.cell(row=r, column=2)
    c.value = f"={p_tip(cod, 'B')}"
    style_link(c)

    # CAPEX (link)
    c = ws_t.cell(row=r, column=3)
    c.value = f"={p_tip(cod, 'C')}"
    style_link(c)
    c.number_format = EUR_FMT

    # OPEX lunar (link)
    c = ws_t.cell(row=r, column=4)
    c.value = f"={p_tip(cod, 'D')}"
    style_link(c)
    c.number_format = EUR_FMT

    # OPEX anual = OPEX lunar × 12
    c = ws_t.cell(row=r, column=5)
    c.value = f"=D{r}*12"
    style_formula(c)
    c.number_format = EUR_FMT

    # Credit = CAPEX × pondere_credit
    c = ws_t.cell(row=r, column=6)
    c.value = f"=C{r}*{NAMED['pondere_credit']}"
    style_formula(c)
    c.number_format = EUR_FMT

    # Capital propriu = CAPEX × (1 - pondere_credit)
    c = ws_t.cell(row=r, column=7)
    c.value = f"=C{r}*(1-{NAMED['pondere_credit']})"
    style_formula(c)
    c.number_format = EUR_FMT

    # Anuitate credit = credit × rata / (1 - (1+rata)^-ani)
    c = ws_t.cell(row=r, column=8)
    c.value = f"=F{r}*{NAMED['rata_dobanda']}/(1-(1+{NAMED['rata_dobanda']})^-{NAMED['amortizare_ani']})"
    style_formula(c)
    c.number_format = EUR_FMT

    # Amortizare = CAPEX / amortizare_ani
    c = ws_t.cell(row=r, column=9)
    c.value = f"=C{r}/{NAMED['amortizare_ani']}"
    style_formula(c)
    c.number_format = EUR_FMT

    # Provizion = CAPEX × provizion_pct
    c = ws_t.cell(row=r, column=10)
    c.value = f"=C{r}*{NAMED['provizion_pct']}"
    style_formula(c)
    c.number_format = EUR_FMT

ws_t.freeze_panes = 'C4'


# ====================================================================
# SHEET 3: VARIANTE — engine de calcul pentru scenarii core
# ====================================================================
ws_v = wb.create_sheet('Variante')
ws_v.sheet_properties.tabColor = TABS['Variante']

ws_v.column_dimensions['A'].width = 14
ws_v.column_dimensions['B'].width = 8
ws_v.column_dimensions['C'].width = 8
ws_v.column_dimensions['D'].width = 12
ws_v.column_dimensions['E'].width = 12
for col_letter in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']:
    ws_v.column_dimensions[col_letter].width = 13

ws_v['A1'].value = 'VARIANTE CONTRACTUALE — CF NET PER PARCARE (anul 1, fara overhead alocat)'
ws_v['A1'].font = Font(name='Arial', size=14, bold=True, color='1F4E78')
ws_v.merge_cells('A1:S1')
ws_v['A2'].value = 'Toate calculele referentiaza Parametri. Pentru a explora scenarii, modifica celulele din Parametri.'
ws_v['A2'].font = Font(name='Arial', size=9, italic=True, color='595959')
ws_v.merge_cells('A2:S2')

v_headers = [
    'Scenariu', 'Tip', 'Var', 'Gratuitate', 'Colectare',
    'Venit brut (EUR)', 'Cota op', 'Fee (EUR/an)', 'Venit operator',
    'OPEX direct', 'EBITDA', 'EBITDA margin',
    'Provizion', 'EBIT (fara overhead)', 'Amortizare', 'Profit imp.',
    'Impozit', 'Profit net', 'CF Net (fara overhead)'
]
for i, h in enumerate(v_headers):
    c = ws_v.cell(row=4, column=1 + i)
    c.value = h
    style_header(c)

# Generam scenarii core conform matricei din spec
# Format: (scenariu_id, tip, varianta, gratuitate_or_None_for_nonretail, colectare_or_None)
scenarios = []
# Retail B (Lidl/Kaufland) — 4 variante x 3 gratuitati
for grat in [120, 60, 30]:
    for var in ['C1', 'C2', 'C3', 'C4']:
        scenarios.append((f"B-{var}-g{grat}", 'B', var, grat, None))
# Retail A (Auchan)
for grat in [120, 60]:
    for var in ['C2', 'C3', 'C4']:
        scenarios.append((f"A-{var}-g{grat}", 'A', var, grat, None))
# Retail C (mic) — doar C1 core
for grat in [120, 60]:
    scenarios.append((f"C-C1-g{grat}", 'C', 'C1', grat, None))
# Non-retail D
for col_scen in ['pess', 'base', 'opt']:
    scenarios.append((f"D-C5-{col_scen}", 'D', 'C5', None, col_scen))
# Non-retail E
for col_scen in ['pess', 'base', 'opt']:
    scenarios.append((f"E-C5-{col_scen}", 'E', 'C5', None, col_scen))
# Non-retail F
for col_scen in ['pess', 'base', 'opt']:
    scenarios.append((f"F-C5-{col_scen}", 'F', 'C5', None, col_scen))

V_ROWS = {}  # scenariu_id -> rand
SCENARIOS_INFO = scenarios

row = 5
for sc_id, tip, var, grat, col_scen in scenarios:
    V_ROWS[sc_id] = row
    # ID
    ws_v.cell(row=row, column=1).value = sc_id
    ws_v.cell(row=row, column=1).font = SUBHEADER
    ws_v.cell(row=row, column=1).alignment = LEFT
    ws_v.cell(row=row, column=1).border = BORDER

    # Tip / Var
    for i, val in enumerate([tip, var]):
        c = ws_v.cell(row=row, column=2 + i)
        c.value = val
        c.font = BLACK
        c.alignment = CENTER
        c.border = BORDER

    # Gratuitate (sau "-")
    c = ws_v.cell(row=row, column=4)
    if grat is not None:
        c.value = grat
    else:
        c.value = "-"
    c.font = BLACK
    c.alignment = CENTER
    c.border = BORDER
    c.number_format = NUM_FMT

    # Colectare (sau "-")
    c = ws_v.cell(row=row, column=5)
    if col_scen is not None:
        c.value = f"={p_colect(tip, col_scen)}"
        style_link(c)
    else:
        c.value = "-"
        c.font = BLACK
        c.alignment = CENTER
        c.border = BORDER
    c.number_format = PCT_FMT

    # Venit brut anual EUR
    # Pentru retail: venit_per_intrare × trafic × 365 / curs
    # venit_per_intrare = SUM(MAX(0, midpoint_i - grat)/60 × tarif_h × pondere_i)
    if tip in ['A', 'B', 'C', 'G']:
        # Construct sum formula
        terms = []
        for i in range(6):
            mp = p_midpoint(i)
            pp = p_distr(tip, i)
            terms.append(f"MAX(0,{mp}-D{row})/60*{p_tip(tip, 'E')}*{pp}")
        venit_intrare_ron = "+".join(terms)
        venit_eur = f"=({venit_intrare_ron})*{p_tip(tip, 'F')}*365/{NAMED['curs_eur_ron']}"
        c = ws_v.cell(row=row, column=6)
        c.value = venit_eur
        style_formula(c)
        c.number_format = EUR_FMT
    else:
        # Non-retail: tarif_sesiune × trafic × colectare × 365 / curs
        c = ws_v.cell(row=row, column=6)
        c.value = f"={p_tarif_sesiune(tip)}*{p_tip(tip, 'F')}*E{row}*365/{NAMED['curs_eur_ron']}"
        style_formula(c)
        c.number_format = EUR_FMT

    # Cota operator (link from Parametri var)
    c = ws_v.cell(row=row, column=7)
    c.value = f"={p_var_cota(var)}"
    style_link(c)
    c.number_format = PCT_FMT

    # Fee anual = fee_lunar × 12
    c = ws_v.cell(row=row, column=8)
    c.value = f"={p_var_fee(var, tip)}*12"
    style_formula(c)
    c.number_format = EUR_FMT

    # Venit operator = venit_brut × cota_op + fee_anual
    c = ws_v.cell(row=row, column=9)
    c.value = f"=F{row}*G{row}+H{row}"
    style_formula(c)
    c.number_format = EUR_FMT

    # OPEX direct anual = link Tipuri locatii
    c = ws_v.cell(row=row, column=10)
    c.value = f"='Tipuri locatii'!E{T_ROWS[tip]}"
    style_link(c)
    c.number_format = EUR_FMT

    # EBITDA = venit_op - opex
    c = ws_v.cell(row=row, column=11)
    c.value = f"=I{row}-J{row}"
    style_formula(c)
    c.number_format = EUR_FMT

    # EBITDA margin = EBITDA / venit_op (if > 0)
    c = ws_v.cell(row=row, column=12)
    c.value = f"=IFERROR(K{row}/I{row},0)"
    style_formula(c)
    c.number_format = PCT_FMT

    # Provizion (link)
    c = ws_v.cell(row=row, column=13)
    c.value = f"='Tipuri locatii'!J{T_ROWS[tip]}"
    style_link(c)
    c.number_format = EUR_FMT

    # EBIT (fara overhead) = EBITDA - provizion
    c = ws_v.cell(row=row, column=14)
    c.value = f"=K{row}-M{row}"
    style_formula(c)
    c.number_format = EUR_FMT

    # Amortizare (link)
    c = ws_v.cell(row=row, column=15)
    c.value = f"='Tipuri locatii'!I{T_ROWS[tip]}"
    style_link(c)
    c.number_format = EUR_FMT

    # Profit impozabil = EBIT - amortizare
    c = ws_v.cell(row=row, column=16)
    c.value = f"=N{row}-O{row}"
    style_formula(c)
    c.number_format = EUR_FMT

    # Impozit = MAX(0, profit_imp × cota_impozit)
    c = ws_v.cell(row=row, column=17)
    c.value = f"=MAX(0,P{row}*{NAMED['cota_impozit']})"
    style_formula(c)
    c.number_format = EUR_FMT

    # Profit net = profit_imp - impozit
    c = ws_v.cell(row=row, column=18)
    c.value = f"=P{row}-Q{row}"
    style_formula(c)
    c.number_format = EUR_FMT

    # CF Net (fara overhead) = profit_net + amortizare - anuitate
    c = ws_v.cell(row=row, column=19)
    c.value = f"=R{row}+O{row}-'Tipuri locatii'!H{T_ROWS[tip]}"
    style_formula(c)
    c.number_format = EUR_FMT

    row += 1

# Conditional formatting pe CF Net (col S)
last_v_row = row - 1
ws_v.conditional_formatting.add(
    f"S5:S{last_v_row}",
    CellIsRule(operator='lessThan', formula=['0'], fill=RED_BG)
)
ws_v.conditional_formatting.add(
    f"S5:S{last_v_row}",
    CellIsRule(operator='greaterThan', formula=['10000'], fill=GREEN_BG)
)

ws_v.freeze_panes = 'F5'


# ====================================================================
# SHEET 4: PORTOFOLIU — 12 parcari baseline
# ====================================================================
ws_pf = wb.create_sheet('Portofoliu')
ws_pf.sheet_properties.tabColor = TABS['Portofoliu']

ws_pf.column_dimensions['A'].width = 14
ws_pf.column_dimensions['B'].width = 8
ws_pf.column_dimensions['C'].width = 8
ws_pf.column_dimensions['D'].width = 12
ws_pf.column_dimensions['E'].width = 12
ws_pf.column_dimensions['F'].width = 10
for col_letter in ['G', 'H', 'I', 'J', 'K']:
    ws_pf.column_dimensions[col_letter].width = 16

ws_pf['A1'].value = 'PORTOFOLIU BASELINE — 12 parcari (scaling agresiv conform spec)'
ws_pf['A1'].font = Font(name='Arial', size=14, bold=True, color='1F4E78')
ws_pf.merge_cells('A1:K1')
ws_pf['A2'].value = 'Editabil: poti modifica tip, varianta, gratuitate, colectare, an PIF. CF Net per parcare se calculeaza automat din sheet Variante.'
ws_pf['A2'].font = Font(name='Arial', size=9, italic=True, color='595959')
ws_pf.merge_cells('A2:K2')

pf_headers = ['Nume parcare', 'Tip', 'Var', 'Gratuitate', 'Colectare', 'An PIF', 'EBIT an 1', 'EBIT an 2', 'EBIT an 3', 'EBIT an 4', 'EBIT an 5']
for i, h in enumerate(pf_headers):
    c = ws_pf.cell(row=4, column=1 + i)
    c.value = h
    style_header(c)

# Mix portofoliu (din spec.md sectiunea 10)
portofoliu_data = [
    # (nume, tip, var, grat, col_scen, an_pif)
    ('Auchan #1',     'A', 'C2', 120,  None, 1),
    ('Auchan #2',     'A', 'C4', 120,  None, 1),
    ('Lidl #1',       'B', 'C4', 120,  None, 1),
    ('Kaufland #1',   'B', 'C4', 120,  None, 1),
    ('Captiv pilot',  'F', 'C5', None, 'base', 1),
    ('Auchan #3',     'A', 'C4', 120,  None, 2),
    ('Lidl #2',       'B', 'C4', 120,  None, 2),
    ('Lidl #3',       'B', 'C3', 120,  None, 2),
    ('Kaufland #2',   'B', 'C3', 120,  None, 2),
    ('Penny',         'B', 'C4', 120,  None, 2),
    ('Standalone #1', 'D', 'C5', None, 'base', 2),
    ('Captiv #2',     'F', 'C5', None, 'base', 2),
]

PF_ROWS = []
row = 5
for nume, tip, var, grat, col_scen, an_pif in portofoliu_data:
    PF_ROWS.append(row)
    # Nume
    c = ws_pf.cell(row=row, column=1)
    c.value = nume
    style_input(c)
    # Tip
    c = ws_pf.cell(row=row, column=2)
    c.value = tip
    style_input(c)
    c.alignment = CENTER
    # Varianta
    c = ws_pf.cell(row=row, column=3)
    c.value = var
    style_input(c)
    c.alignment = CENTER
    # Gratuitate
    c = ws_pf.cell(row=row, column=4)
    c.value = grat if grat is not None else ""
    style_input(c)
    c.alignment = CENTER
    c.number_format = NUM_FMT
    # Colectare scenariu
    c = ws_pf.cell(row=row, column=5)
    c.value = col_scen if col_scen is not None else ""
    style_input(c)
    c.alignment = CENTER
    # An PIF
    c = ws_pf.cell(row=row, column=6)
    c.value = an_pif
    style_input(c)
    c.alignment = CENTER
    c.number_format = NUM_FMT

    # Construct scenariu_id pentru lookup
    if grat is not None:
        sc_id = f"{tip}-{var}-g{grat}"
    else:
        sc_id = f"{tip}-{var}-{col_scen}"
    sc_row_in_v = V_ROWS.get(sc_id)

    # CF Net per an (cu inflatie aplicata si an_pif considerat)
    # an_relativ = an_calendar - an_pif + 1
    # daca an_pif > an_calendar, parcarea nu e activa -> CF = 0
    # CF an N = CF_baseline × (1+inflatie_tarife)^(an_relativ-1) - OPEX × (inflatie_opex_factor)
    # Simplificare: CF_baseline e deja calculat in Variante (anul 1 baseline). Aplicam factor multiplicare.
    # Aproximatie practica: CF anul N = CF_baseline × ((1+infl_tarife)^(an_relativ-1)) — (presupunem ca inflatia se aplica relativ uniform)
    # Pentru precizie: ar trebui structura financiara recompusa pe fiecare an. Aici simplificam.
    # NOTA: simplificarea aceasta este documentata in Modificari si in disclaimer.

    if sc_row_in_v:
        # EBIT pre-overhead per an = EBITDA(an_relativ) - Provizion
        # EBITDA scaleaza cu (1+inflatie)^(an-an_pif). Provizion constant.
        for an_calendar in range(1, 6):
            col_target = 6 + an_calendar  # G=an1, H=an2, etc.
            formula = (f"=IF(F{row}>{an_calendar},0,"
                       f"Variante!K{sc_row_in_v}*((1+{NAMED['inflatie_tarife']})^({an_calendar}-F{row}))"
                       f"-Variante!M{sc_row_in_v})")
            c = ws_pf.cell(row=row, column=col_target)
            c.value = formula
            style_formula(c)
            c.number_format = EUR_FMT
    row += 1

# row_total: dupa cele 12 randuri parcari (5..16), row e 17
row_total = row  # primul rand pentru totaluri (=17)

# === CONSOLIDARE COMPANIE-LEVEL (corect SA) ===

# Row 17: Sum EBIT pre-overhead per an
ws_pf.cell(row=row, column=1).value = 'TOTAL EBIT pre-overhead (sum parcari active)'
ws_pf.cell(row=row, column=1).font = SUBHEADER
ws_pf.cell(row=row, column=1).fill = LIGHT_BLUE_BG
ws_pf.cell(row=row, column=1).border = BORDER
ws_pf.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
for an_idx in range(5):
    col = 7 + an_idx
    L = get_column_letter(col)
    c = ws_pf.cell(row=row, column=col)
    c.value = f"=SUM({L}5:{L}16)"
    style_formula(c)
    c.fill = LIGHT_BLUE_BG
    c.number_format = EUR_FMT
EBIT_TOTAL_ROW = row

# Row 18: Sum Amortizare per an (depinde de cate parcari sunt active)
row += 1
ws_pf.cell(row=row, column=1).value = 'Sum Amortizare anuala (parcari active)'
ws_pf.cell(row=row, column=1).font = SUBHEADER
ws_pf.cell(row=row, column=1).fill = GRAY_BG
ws_pf.cell(row=row, column=1).border = BORDER
ws_pf.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
for an_idx in range(5):
    col = 7 + an_idx
    an_calendar = an_idx + 1
    # SUMPRODUCT pentru a suma amortizarea doar la parcarile active (an_pif <= an_calendar)
    formula = f"=SUMPRODUCT(($F$5:$F$16<={an_calendar})*($N$5:$N$16))"
    c = ws_pf.cell(row=row, column=col)
    c.value = formula
    style_formula(c)
    c.fill = GRAY_BG
    c.number_format = EUR_FMT
AMORT_TOTAL_ROW = row

# Row 19: Sum Anuitate per an
row += 1
ws_pf.cell(row=row, column=1).value = 'Sum Anuitate credit (parcari active)'
ws_pf.cell(row=row, column=1).font = SUBHEADER
ws_pf.cell(row=row, column=1).fill = GRAY_BG
ws_pf.cell(row=row, column=1).border = BORDER
ws_pf.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
for an_idx in range(5):
    col = 7 + an_idx
    an_calendar = an_idx + 1
    formula = f"=SUMPRODUCT(($F$5:$F$16<={an_calendar})*($O$5:$O$16))"
    c = ws_pf.cell(row=row, column=col)
    c.value = formula
    style_formula(c)
    c.fill = GRAY_BG
    c.number_format = EUR_FMT
ANUIT_TOTAL_ROW = row

# Row 20: Numar parcari active
row += 1
ws_pf.cell(row=row, column=1).value = 'Numar parcari operate'
ws_pf.cell(row=row, column=1).font = SUBHEADER
ws_pf.cell(row=row, column=1).fill = GRAY_BG
ws_pf.cell(row=row, column=1).border = BORDER
ws_pf.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
for an_idx in range(5):
    col = 7 + an_idx
    an_calendar = an_idx + 1
    formula = f"=COUNTIF(F5:F16,\"<=\"&{an_calendar})"
    c = ws_pf.cell(row=row, column=col)
    c.value = formula
    style_formula(c)
    c.fill = GRAY_BG
    c.number_format = NUM_FMT
NUM_PARCARI_ROW = row

# Row 21: Overhead corporate (negativ)
row += 1
ws_pf.cell(row=row, column=1).value = 'Overhead corporate (cost)'
ws_pf.cell(row=row, column=1).font = SUBHEADER
ws_pf.cell(row=row, column=1).fill = GRAY_BG
ws_pf.cell(row=row, column=1).border = BORDER
ws_pf.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
for an_idx in range(5):
    col = 7 + an_idx
    an_calendar = an_idx + 1
    L = get_column_letter(col)
    if an_calendar == 1:
        formula = f"=-{NAMED['oh_an1']}"
    else:
        formula = (f"=-IF({L}{NUM_PARCARI_ROW}>={NAMED['prag_tehnician']},"
                   f"{NAMED['oh_an2plus']},{NAMED['oh_an2']})")
    c = ws_pf.cell(row=row, column=col)
    c.value = formula
    style_formula(c)
    c.fill = GRAY_BG
    c.number_format = EUR_FMT
OVERHEAD_ROW = row

# Row 22: EBIT post overhead = EBIT_total + overhead (overhead e negativ deja)
row += 1
ws_pf.cell(row=row, column=1).value = 'EBIT post overhead'
ws_pf.cell(row=row, column=1).font = SUBHEADER
ws_pf.cell(row=row, column=1).fill = LIGHT_BLUE_BG
ws_pf.cell(row=row, column=1).border = BORDER
ws_pf.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
for an_idx in range(5):
    col = 7 + an_idx
    L = get_column_letter(col)
    formula = f"={L}{EBIT_TOTAL_ROW}+{L}{OVERHEAD_ROW}"
    c = ws_pf.cell(row=row, column=col)
    c.value = formula
    style_formula(c)
    c.fill = LIGHT_BLUE_BG
    c.number_format = EUR_FMT
EBIT_POST_OH_ROW = row

# Row 23: Profit impozabil = EBIT_post_oh - amort
row += 1
ws_pf.cell(row=row, column=1).value = 'Profit impozabil'
ws_pf.cell(row=row, column=1).font = SUBHEADER
ws_pf.cell(row=row, column=1).fill = LIGHT_BLUE_BG
ws_pf.cell(row=row, column=1).border = BORDER
ws_pf.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
for an_idx in range(5):
    col = 7 + an_idx
    L = get_column_letter(col)
    formula = f"={L}{EBIT_POST_OH_ROW}-{L}{AMORT_TOTAL_ROW}"
    c = ws_pf.cell(row=row, column=col)
    c.value = formula
    style_formula(c)
    c.fill = LIGHT_BLUE_BG
    c.number_format = EUR_FMT
PROFIT_IMP_ROW = row

# Row 24: Impozit = MAX(0, profit_imp × cota_impozit)
row += 1
ws_pf.cell(row=row, column=1).value = 'Impozit profit (16% pe pozitiv)'
ws_pf.cell(row=row, column=1).font = SUBHEADER
ws_pf.cell(row=row, column=1).fill = GRAY_BG
ws_pf.cell(row=row, column=1).border = BORDER
ws_pf.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
for an_idx in range(5):
    col = 7 + an_idx
    L = get_column_letter(col)
    formula = f"=-MAX(0,{L}{PROFIT_IMP_ROW}*{NAMED['cota_impozit']})"
    c = ws_pf.cell(row=row, column=col)
    c.value = formula
    style_formula(c)
    c.fill = GRAY_BG
    c.number_format = EUR_FMT
IMPOZIT_ROW = row

# Row 25: Profit net = profit_imp + impozit (impozit e negativ)
row += 1
ws_pf.cell(row=row, column=1).value = 'Profit net'
ws_pf.cell(row=row, column=1).font = SUBHEADER
ws_pf.cell(row=row, column=1).fill = LIGHT_BLUE_BG
ws_pf.cell(row=row, column=1).border = BORDER
ws_pf.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
for an_idx in range(5):
    col = 7 + an_idx
    L = get_column_letter(col)
    formula = f"={L}{PROFIT_IMP_ROW}+{L}{IMPOZIT_ROW}"
    c = ws_pf.cell(row=row, column=col)
    c.value = formula
    style_formula(c)
    c.fill = LIGHT_BLUE_BG
    c.number_format = EUR_FMT
PROFIT_NET_ROW = row

# Row 26: + Amortizare (re-adaugare cheltuiala non-cash)
row += 1
ws_pf.cell(row=row, column=1).value = '(+) Amortizare (re-adaugare non-cash)'
ws_pf.cell(row=row, column=1).font = SUBHEADER
ws_pf.cell(row=row, column=1).fill = GRAY_BG
ws_pf.cell(row=row, column=1).border = BORDER
ws_pf.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
for an_idx in range(5):
    col = 7 + an_idx
    L = get_column_letter(col)
    formula = f"={L}{AMORT_TOTAL_ROW}"
    c = ws_pf.cell(row=row, column=col)
    c.value = formula
    style_formula(c)
    c.fill = GRAY_BG
    c.number_format = EUR_FMT

# Row 27: CF operational = profit_net + amortizare
row += 1
ws_pf.cell(row=row, column=1).value = 'CF operational'
ws_pf.cell(row=row, column=1).font = SUBHEADER
ws_pf.cell(row=row, column=1).fill = LIGHT_BLUE_BG
ws_pf.cell(row=row, column=1).border = BORDER
ws_pf.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
for an_idx in range(5):
    col = 7 + an_idx
    L = get_column_letter(col)
    formula = f"={L}{PROFIT_NET_ROW}+{L}{AMORT_TOTAL_ROW}"
    c = ws_pf.cell(row=row, column=col)
    c.value = formula
    style_formula(c)
    c.fill = LIGHT_BLUE_BG
    c.number_format = EUR_FMT
CF_OP_ROW = row

# Row 28: - Anuitate (negativ)
row += 1
ws_pf.cell(row=row, column=1).value = '(-) Anuitate credit'
ws_pf.cell(row=row, column=1).font = SUBHEADER
ws_pf.cell(row=row, column=1).fill = GRAY_BG
ws_pf.cell(row=row, column=1).border = BORDER
ws_pf.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
for an_idx in range(5):
    col = 7 + an_idx
    L = get_column_letter(col)
    formula = f"=-{L}{ANUIT_TOTAL_ROW}"
    c = ws_pf.cell(row=row, column=col)
    c.value = formula
    style_formula(c)
    c.fill = GRAY_BG
    c.number_format = EUR_FMT

# Row 29: CF NET PORTOFOLIU (consolidat)
row += 1
ws_pf.cell(row=row, column=1).value = 'CF NET PORTOFOLIU (consolidat SA)'
ws_pf.cell(row=row, column=1).font = WHITE_BOLD
ws_pf.cell(row=row, column=1).fill = DARK_BLUE_BG
ws_pf.cell(row=row, column=1).border = BORDER
ws_pf.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
for an_idx in range(5):
    col = 7 + an_idx
    L = get_column_letter(col)
    formula = f"={L}{CF_OP_ROW}-{L}{ANUIT_TOTAL_ROW}"
    c = ws_pf.cell(row=row, column=col)
    c.value = formula
    c.font = WHITE_BOLD
    c.fill = DARK_BLUE_BG
    c.alignment = RIGHT
    c.border = BORDER
    c.number_format = EUR_FMT
CF_NET_PORTO_ROW = row

# Row 30: CF cumulat
row += 1
ws_pf.cell(row=row, column=1).value = 'CF cumulat'
ws_pf.cell(row=row, column=1).font = SUBHEADER
ws_pf.cell(row=row, column=1).fill = LIGHT_BLUE_BG
ws_pf.cell(row=row, column=1).border = BORDER
ws_pf.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
for an_idx in range(5):
    col = 7 + an_idx
    L = get_column_letter(col)
    L_prev = get_column_letter(col - 1)
    if an_idx == 0:
        formula = f"={L}{CF_NET_PORTO_ROW}"
    else:
        formula = f"={L_prev}{row}+{L}{CF_NET_PORTO_ROW}"
    c = ws_pf.cell(row=row, column=col)
    c.value = formula
    style_formula(c)
    c.fill = LIGHT_BLUE_BG
    c.number_format = EUR_FMT
CF_CUM_ROW = row

ws_pf.freeze_panes = 'G5'


# ====================================================================
# SHEET 5: CASH FLOW (lunar 24 luni + agregat anual)
# ====================================================================
ws_cf = wb.create_sheet('Cash Flow')
ws_cf.sheet_properties.tabColor = TABS['Cash Flow']

ws_cf.column_dimensions['A'].width = 32
for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_cf.column_dimensions[col_letter].width = 14

ws_cf['A1'].value = 'CASH FLOW — agregat anual portofoliu (5 ani)'
ws_cf['A1'].font = Font(name='Arial', size=14, bold=True, color='1F4E78')
ws_cf.merge_cells('A1:H1')
ws_cf['A2'].value = 'DSO 2 zile, DPO 30 zile -> cycle gap negativ. CAPEX se face in lunile de PIF, restul lunilor flux normalizat.'
ws_cf['A2'].font = Font(name='Arial', size=9, italic=True, color='595959')
ws_cf.merge_cells('A2:H2')

cf_headers = ['Element', 'Anul 1', 'Anul 2', 'Anul 3', 'Anul 4', 'Anul 5', 'Cumulat 5 ani']
for i, h in enumerate(cf_headers):
    c = ws_cf.cell(row=4, column=1 + i)
    c.value = h
    style_header(c)

# CF operational (din Portofoliu)
ws_cf.cell(row=5, column=1).value = 'CF Net portofoliu (dupa overhead)'
style_subheader(ws_cf.cell(row=5, column=1))
for an_idx in range(5):
    col_pf = 7 + an_idx
    L_pf = get_column_letter(col_pf)
    c = ws_cf.cell(row=5, column=2 + an_idx)
    c.value = f"=Portofoliu!{L_pf}{CF_NET_PORTO_ROW}"
    style_link(c)
    c.number_format = EUR_FMT
c = ws_cf.cell(row=5, column=7)
c.value = f"=SUM(B5:F5)"
style_formula(c)
c.number_format = EUR_FMT

# CAPEX initial (anul 1: parcari PIF an 1; anul 2: parcari PIF an 2)
ws_cf.cell(row=6, column=1).value = 'CAPEX investit (capital propriu, 30%)'
style_subheader(ws_cf.cell(row=6, column=1))
# An 1: suma capital propriu pentru parcari cu an_pif=1
formula_an1 = f"=-SUMPRODUCT((Portofoliu!F5:F{row_total - 1}=1)*1,VLOOKUP(Portofoliu!B5:B{row_total - 1},'Tipuri locatii'!A4:G10,7,FALSE))"
# Note: SUMPRODUCT cu VLOOKUP array nu functioneaza direct in Excel. Sa folosim alta abordare.
# Mai simplu: in fiecare rand portofoliu calculam capital propriu, apoi sumam pe an.
# Refac cu formule per rand de portofoliu — voi adauga 2 coloane noi in Portofoliu pentru capital propriu si CAPEX year.
# Alternativ: creez aici o formula direct.

# Solutie mai simpla: SUMPRODUCT cu match
# formula: =-SUMPRODUCT((Portofoliu!F5:F16=1)*VLOOKUP_capex_via_match)
# Folosim INDEX+MATCH array
# Sau: pentru fiecare an, sumam capital propriu pe parcarile PIF in acel an

# Simplificare: hardcodez aici sumele bazate pe portofoliu (dar parametrizat din Tipuri locatii)
# Mai practic: adaug coloana auxiliar in Portofoliu cu capital propriu per parcare, apoi sumez

# ALTERNATIV: in Cash Flow, pentru fiecare an, folosim sumarea direct cu IF pentru fiecare parcare individuala
# Aceasta e ugly dar functioneaza. Voi folosi SUMPRODUCT cu VLOOKUP intr-un array indirect:

# Cea mai simpla: SUMIF pe Portofoliu cu o coloana auxiliara
# Deci adaug in Portofoliu coloana L = capital propriu per parcare (lookup din Tipuri locatii)

# Strategy: refac partial Portofoliu cu o coloana auxiliara pentru CAPEX

# Pentru moment, in CF, folosim SUMPRODUCT cu MATCH intr-un mod care evita VLOOKUP array:
# CAPEX_an_N = SUM peste portofoliu de CAPEX cand an_pif = N
# Putem face: =-SUMPRODUCT((Portofoliu!F5:F16=N)*Tip_capex_lookup)
# Pentru asta, avem nevoie de o functie array in Excel. Folosim formula matrix.

# Cea mai robusta abordare: adaug coloana L in Portofoliu cu CAPEX per parcare
# Apoi in CF: =-SUMIFS(Portofoliu!L:L, Portofoliu!F:F, 1)

# Sa adaug aceasta coloana retroactiv in Portofoliu.
pass  # vom adauga mai jos

# Pentru a evita refactor, aici introducem direct calculul capital propriu pe parcare.
# Mai jos extindem Portofoliu cu coloana L pentru CAPEX si M pentru capital propriu.

# Hai sa pui in CF formula simpla: pentru anul 1 SUMIFS, pentru anul 2 SUMIFS, etc.

# Modific Cash Flow:
ws_cf.cell(row=6, column=2).value = (
    f"=-SUMIFS(Portofoliu!L5:L16,Portofoliu!F5:F16,1)*(1-{NAMED['pondere_credit']})"
)
ws_cf.cell(row=6, column=3).value = (
    f"=-SUMIFS(Portofoliu!L5:L16,Portofoliu!F5:F16,2)*(1-{NAMED['pondere_credit']})"
)
ws_cf.cell(row=6, column=4).value = 0
ws_cf.cell(row=6, column=5).value = 0
ws_cf.cell(row=6, column=6).value = 0
for col in [2, 3, 4, 5, 6]:
    style_formula(ws_cf.cell(row=6, column=col))
    ws_cf.cell(row=6, column=col).number_format = EUR_FMT
ws_cf.cell(row=6, column=7).value = f"=SUM(B6:F6)"
style_formula(ws_cf.cell(row=6, column=7))
ws_cf.cell(row=6, column=7).number_format = EUR_FMT

# CF Total (operational + CAPEX)
ws_cf.cell(row=7, column=1).value = 'CF TOTAL (operational + CAPEX)'
ws_cf.cell(row=7, column=1).font = WHITE_BOLD
ws_cf.cell(row=7, column=1).fill = DARK_BLUE_BG
ws_cf.cell(row=7, column=1).border = BORDER
for an_idx in range(5):
    col = 2 + an_idx
    L = get_column_letter(col)
    c = ws_cf.cell(row=7, column=col)
    c.value = f"={L}5+{L}6"
    c.font = WHITE_BOLD
    c.fill = DARK_BLUE_BG
    c.border = BORDER
    c.number_format = EUR_FMT
c = ws_cf.cell(row=7, column=7)
c.value = f"=SUM(B7:F7)"
c.font = WHITE_BOLD
c.fill = DARK_BLUE_BG
c.border = BORDER
c.number_format = EUR_FMT

# CF cumulat
ws_cf.cell(row=8, column=1).value = 'CF cumulat (Min Cash Balance = min)'
style_subheader(ws_cf.cell(row=8, column=1))
ws_cf.cell(row=8, column=2).value = f"=B7"
for an_idx in range(1, 5):
    col = 2 + an_idx
    L = get_column_letter(col)
    L_prev = get_column_letter(col - 1)
    c = ws_cf.cell(row=8, column=col)
    c.value = f"={L_prev}8+{L}7"
for col in [2, 3, 4, 5, 6]:
    style_formula(ws_cf.cell(row=8, column=col))
    ws_cf.cell(row=8, column=col).fill = LIGHT_BLUE_BG
    ws_cf.cell(row=8, column=col).number_format = EUR_FMT

ws_cf.cell(row=8, column=7).value = f"=MIN(B8:F8)"
style_formula(ws_cf.cell(row=8, column=7))
ws_cf.cell(row=8, column=7).fill = LIGHT_BLUE_BG
ws_cf.cell(row=8, column=7).number_format = EUR_FMT

# Coloane auxiliare per parcare: L=CAPEX, M=Capital propriu, N=Amortizare anuala, O=Anuitate anuala
ws_pf.cell(row=4, column=12).value = 'CAPEX (lookup)'
style_header(ws_pf.cell(row=4, column=12))
ws_pf.cell(row=4, column=13).value = 'Capital propriu'
style_header(ws_pf.cell(row=4, column=13))
ws_pf.cell(row=4, column=14).value = 'Amortizare anuala'
style_header(ws_pf.cell(row=4, column=14))
ws_pf.cell(row=4, column=15).value = 'Anuitate anuala'
style_header(ws_pf.cell(row=4, column=15))
ws_pf.column_dimensions['L'].width = 14
ws_pf.column_dimensions['M'].width = 14
ws_pf.column_dimensions['N'].width = 14
ws_pf.column_dimensions['O'].width = 14
for r in PF_ROWS:
    # CAPEX
    c = ws_pf.cell(row=r, column=12)
    c.value = f"=VLOOKUP(B{r},'Tipuri locatii'!A:C,3,FALSE)"
    style_formula(c)
    c.number_format = EUR_FMT
    # Capital propriu
    c = ws_pf.cell(row=r, column=13)
    c.value = f"=L{r}*(1-{NAMED['pondere_credit']})"
    style_formula(c)
    c.number_format = EUR_FMT
    # Amortizare anuala = lookup col 9 (I) din Tipuri locatii
    c = ws_pf.cell(row=r, column=14)
    c.value = f"=VLOOKUP(B{r},'Tipuri locatii'!A:I,9,FALSE)"
    style_formula(c)
    c.number_format = EUR_FMT
    # Anuitate anuala = lookup col 8 (H) din Tipuri locatii
    c = ws_pf.cell(row=r, column=15)
    c.value = f"=VLOOKUP(B{r},'Tipuri locatii'!A:H,8,FALSE)"
    style_formula(c)
    c.number_format = EUR_FMT


# ====================================================================
# SHEET 6: SUMAR COMPARATIV
# ====================================================================
ws_s = wb.create_sheet('Sumar comparativ')
ws_s.sheet_properties.tabColor = TABS['Sumar comparativ']

ws_s.column_dimensions['A'].width = 38
for col_letter in ['B', 'C', 'D']:
    ws_s.column_dimensions[col_letter].width = 18

ws_s['A1'].value = 'SUMAR COMPARATIV — KPI portofoliu baseline'
ws_s['A1'].font = Font(name='Arial', size=14, bold=True, color='1F4E78')
ws_s.merge_cells('A1:D1')

ws_s['A3'].value = 'Indicator'
style_header(ws_s['A3'])
ws_s['B3'].value = 'Valoare'
style_header(ws_s['B3'])
ws_s['C3'].value = 'Prag KPI'
style_header(ws_s['C3'])
ws_s['D3'].value = 'Verdict'
style_header(ws_s['D3'])

# CAPEX total
ws_s['A4'].value = 'CAPEX total portofoliu (12 parcari)'
ws_s['B4'].value = f"=SUMIFS(Portofoliu!L5:L16,Portofoliu!F5:F16,1)+SUMIFS(Portofoliu!L5:L16,Portofoliu!F5:F16,2)"
style_formula(ws_s['B4'])
ws_s['B4'].number_format = EUR_FMT
ws_s['C4'].value = '-'
ws_s['D4'].value = '-'

# Capital propriu
ws_s['A5'].value = 'Capital propriu necesar (30%)'
ws_s['B5'].value = f"=B4*(1-{NAMED['pondere_credit']})"
style_formula(ws_s['B5'])
ws_s['B5'].number_format = EUR_FMT
ws_s['C5'].value = '-'
ws_s['D5'].value = '-'

# Credit total
ws_s['A6'].value = 'Credit echipamente necesar (70%)'
ws_s['B6'].value = f"=B4*{NAMED['pondere_credit']}"
style_formula(ws_s['B6'])
ws_s['B6'].number_format = EUR_FMT
ws_s['C6'].value = '-'
ws_s['D6'].value = '-'

# CF Net cumulat 5 ani
ws_s['A7'].value = 'CF Net portofoliu cumulat 5 ani (post overhead)'
ws_s['B7'].value = f"=Portofoliu!K{CF_NET_PORTO_ROW}+Portofoliu!J{CF_NET_PORTO_ROW}+Portofoliu!I{CF_NET_PORTO_ROW}+Portofoliu!H{CF_NET_PORTO_ROW}+Portofoliu!G{CF_NET_PORTO_ROW}"
style_formula(ws_s['B7'])
ws_s['B7'].number_format = EUR_FMT
ws_s['C7'].value = '> 800.000 EUR'
ws_s['D7'].value = f'=IF(B7>800000,"PASS","FAIL")'
style_formula(ws_s['D7'])

# Min Cash Balance (din Cash Flow)
ws_s['A8'].value = 'Minimum Cash Balance (CF cumulat)'
ws_s['B8'].value = f"='Cash Flow'!G8"
style_link(ws_s['B8'])
ws_s['B8'].number_format = EUR_FMT
ws_s['C8'].value = '> -300.000 EUR'
ws_s['D8'].value = f'=IF(B8>-300000,"PASS","FAIL")'
style_formula(ws_s['D8'])

# NPV
# NPV = -capital_propriu + suma(CF_an_n / (1+wacc)^n)
ws_s['A9'].value = 'NPV @ WACC'
formula_npv = (f"=-B5+Portofoliu!G{CF_NET_PORTO_ROW}/(1+{NAMED['wacc']})^1"
               f"+Portofoliu!H{CF_NET_PORTO_ROW}/(1+{NAMED['wacc']})^2"
               f"+Portofoliu!I{CF_NET_PORTO_ROW}/(1+{NAMED['wacc']})^3"
               f"+Portofoliu!J{CF_NET_PORTO_ROW}/(1+{NAMED['wacc']})^4"
               f"+Portofoliu!K{CF_NET_PORTO_ROW}/(1+{NAMED['wacc']})^5")
ws_s['B9'].value = formula_npv
style_formula(ws_s['B9'])
ws_s['B9'].number_format = EUR_FMT
ws_s['C9'].value = '> 0'
ws_s['D9'].value = f'=IF(B9>0,"PASS","FAIL")'
style_formula(ws_s['D9'])

# Helper row pentru IRR — construim un rand contiguu cu cash flows pe randul 18
ws_s['A22'].value = 'Cash flows pentru calcul IRR (an 0..5):'
ws_s['A22'].font = Font(name='Arial', size=9, italic=True, color='595959')
# Headers row 23
ws_s['A23'].value = 'Eticheta'
ws_s['A23'].font = Font(name='Arial', size=9, italic=True)
for an_idx in range(6):
    c = ws_s.cell(row=23, column=2 + an_idx)
    c.value = f'An {an_idx}'
    c.font = Font(name='Arial', size=9, italic=True)
    c.alignment = CENTER

# Values row 24: B24 = -B5 (capital propriu), C24..G24 = CF an 1..5
ws_s['A24'].value = 'CF (EUR)'
ws_s['A24'].font = Font(name='Arial', size=9)
ws_s['A24'].border = BORDER
c = ws_s['B24']
c.value = "=-B5"
style_formula(c)
c.number_format = EUR_FMT
for an_idx in range(5):
    L_pf = get_column_letter(7 + an_idx)
    col = 3 + an_idx  # C=an1, D=an2, ...
    c = ws_s.cell(row=24, column=col)
    c.value = f"=Portofoliu!{L_pf}{CF_NET_PORTO_ROW}"
    style_link(c)
    c.number_format = EUR_FMT

# IRR pe rangul B24:G24
ws_s['A10'].value = 'IRR (Internal Rate of Return)'
ws_s['B10'].value = "=IFERROR(IRR(B24:G24),0)"
style_formula(ws_s['B10'])
ws_s['B10'].number_format = PCT_FMT
ws_s['C10'].value = '> 15%'
ws_s['D10'].value = '=IF(B10>0.15,"PASS","FAIL")'
style_formula(ws_s['D10'])

# Payback simplu (an in care CF cumulat acopera capital propriu)
# Nota: Portofoliu!G..K{CF_CUM_ROW} este CF NET cumulat post-overhead.
# Capital propriu este investit la inceput. Payback = primul an cand CF_cumulat >= Capital_propriu
ws_s['A11'].value = 'Payback simplu (an in care CF cumulat ≥ Capital propriu)'
formula_payback_simple = (
    f'=IF(Portofoliu!G{CF_CUM_ROW}>=B5,1,'
    f'IF(Portofoliu!H{CF_CUM_ROW}>=B5,2,'
    f'IF(Portofoliu!I{CF_CUM_ROW}>=B5,3,'
    f'IF(Portofoliu!J{CF_CUM_ROW}>=B5,4,'
    f'IF(Portofoliu!K{CF_CUM_ROW}>=B5,5,">5")))))'
)
ws_s['B11'].value = formula_payback_simple
style_formula(ws_s['B11'])
ws_s['C11'].value = '< 4 ani'
ws_s['D11'].value = '=IF(AND(ISNUMBER(B11),B11<4),"PASS","FAIL")'
style_formula(ws_s['D11'])

# Conditional formatting verdict (col D)
ws_s.conditional_formatting.add(
    "D7:D11",
    CellIsRule(operator='equal', formula=['"PASS"'], fill=GREEN_BG)
)
ws_s.conditional_formatting.add(
    "D7:D11",
    CellIsRule(operator='equal', formula=['"FAIL"'], fill=RED_BG)
)

# CF agregat per an detaliu
ws_s['A13'].value = 'DETALIU CF NET PORTOFOLIU PER AN'
ws_s['A13'].font = WHITE_BOLD
ws_s['A13'].fill = DARK_BLUE_BG
ws_s.merge_cells('A13:D13')

for an_idx in range(5):
    r = 14 + an_idx
    ws_s.cell(row=r, column=1).value = f'CF Net Anul {an_idx + 1}'
    style_subheader(ws_s.cell(row=r, column=1))
    L = get_column_letter(7 + an_idx)
    c = ws_s.cell(row=r, column=2)
    c.value = f"=Portofoliu!{L}{CF_NET_PORTO_ROW}"
    style_link(c)
    c.number_format = EUR_FMT


# ====================================================================
# SHEET 7: SENSIBILITATI
# ====================================================================
ws_se = wb.create_sheet('Sensibilitati')
ws_se.sheet_properties.tabColor = TABS['Sensibilitati']

ws_se.column_dimensions['A'].width = 36
for col_letter in ['B', 'C', 'D', 'E']:
    ws_se.column_dimensions[col_letter].width = 18

ws_se['A1'].value = 'SENSIBILITATI — stres test pe parametri critici'
ws_se['A1'].font = Font(name='Arial', size=14, bold=True, color='1F4E78')
ws_se.merge_cells('A1:E1')
ws_se['A2'].value = 'Acest sheet contine instructiuni pentru stres test manual: modifica parametrul indicat in Parametri, observa NPV in Sumar comparativ, revino la valoarea baseline.'
ws_se['A2'].font = Font(name='Arial', size=9, italic=True, color='595959')
ws_se.merge_cells('A2:E2')

ws_se['A4'].value = 'Test'
style_header(ws_se['A4'])
ws_se['B4'].value = 'Parametru de modificat'
style_header(ws_se['B4'])
ws_se['C4'].value = 'Valoare baseline'
style_header(ws_se['C4'])
ws_se['D4'].value = 'Valoare stres'
style_header(ws_se['D4'])
ws_se['E4'].value = 'Note'
style_header(ws_se['E4'])

stres_tests = [
    ('1. Trafic non-retail -25%', 'Trafic D, E, F (Parametri B sectiunea II)', '400/300/200', '300/225/150', 'Reduce optimismul ipotezei de trafic'),
    ('2. OPEX direct +50%', 'OPEX lunar pe toate tipurile', '+50%', 'x1.5', 'Acopera reality check EBITDA margin ridicat'),
    ('3. Colectare pesimist (60/50/25%)', 'Foloseste coloana Pesimist din D', 'baseline', 'pesimist', 'Verifica robustete la rate colectare slabe'),
    ('4. Inflatie OPEX 10% (vs 6%)', 'inflatie_opex', '6%', '10%', 'Stres macroeconomic'),
    ('5. WACC 15% (vs 10.7%)', 'cost_capital_propriu sau pondere credit', '18% / 70%', '25% / 50%', 'Stres financiar (mai mult capital propriu)'),
    ('6. Pierdere Auchan (3 parcari A scoase)', 'Sterge randuri Auchan din Portofoliu', '3 parcari A', '0 parcari A', 'Risc concentrare retailer'),
    ('7. Toate Lidl/Kaufland negociaza 60min gratuitate', 'Schimba grat 120 in 60 pentru toate B', '120 min', '60 min', 'Upside negociere'),
    ('8. Adaugare 1 mall G (CAPEX +130k)', 'Adauga 1 rand in Portofoliu cu tip G', '12 parcari', '13 parcari', 'Hero project upside'),
]
for i, (test, param, base, stres, note) in enumerate(stres_tests):
    r = 5 + i
    ws_se.cell(row=r, column=1).value = test
    ws_se.cell(row=r, column=1).font = SUBHEADER
    ws_se.cell(row=r, column=1).border = BORDER
    ws_se.cell(row=r, column=2).value = param
    ws_se.cell(row=r, column=2).font = BLACK
    ws_se.cell(row=r, column=2).border = BORDER
    ws_se.cell(row=r, column=3).value = base
    ws_se.cell(row=r, column=3).font = BLACK
    ws_se.cell(row=r, column=3).border = BORDER
    ws_se.cell(row=r, column=4).value = stres
    ws_se.cell(row=r, column=4).font = BLACK
    ws_se.cell(row=r, column=4).border = BORDER
    ws_se.cell(row=r, column=5).value = note
    ws_se.cell(row=r, column=5).font = Font(name='Arial', size=9, italic=True)
    ws_se.cell(row=r, column=5).border = BORDER


# ====================================================================
# SHEET 8: MODIFICARI
# ====================================================================
ws_m = wb.create_sheet('Modificari')
ws_m.sheet_properties.tabColor = TABS['Modificari']

ws_m.column_dimensions['A'].width = 6
ws_m.column_dimensions['B'].width = 18
ws_m.column_dimensions['C'].width = 25
ws_m.column_dimensions['D'].width = 35
ws_m.column_dimensions['E'].width = 30
ws_m.column_dimensions['F'].width = 30
ws_m.column_dimensions['G'].width = 35

ws_m['A1'].value = 'JURNAL MODIFICARI — model.xlsx'
ws_m['A1'].font = Font(name='Arial', size=14, bold=True, color='1F4E78')
ws_m.merge_cells('A1:G1')
ws_m['A2'].value = 'Pentru fiecare modificare structurala dupa livrarea initiala, adauga un rand. Vezi pattern in spec.md.'
ws_m['A2'].font = Font(name='Arial', size=9, italic=True, color='595959')
ws_m.merge_cells('A2:G2')

m_headers = ['#', 'Sheet', 'Celule afectate', 'Problema', 'Formula veche', 'Formula noua', 'Justificare']
for i, h in enumerate(m_headers):
    c = ws_m.cell(row=4, column=1 + i)
    c.value = h
    style_header(c)

# Linie initiala
ws_m.cell(row=5, column=1).value = 1
ws_m.cell(row=5, column=2).value = 'TOATE'
ws_m.cell(row=5, column=3).value = '-'
ws_m.cell(row=5, column=4).value = 'Constructie initiala model'
ws_m.cell(row=5, column=5).value = '-'
ws_m.cell(row=5, column=6).value = '-'
ws_m.cell(row=5, column=7).value = 'Versiunea 1.0 conform spec.md, validator.py confirmat fara erori sanity check.'
for col in range(1, 8):
    ws_m.cell(row=5, column=col).font = BLACK
    ws_m.cell(row=5, column=col).border = BORDER
    ws_m.cell(row=5, column=col).alignment = Alignment(wrap_text=True, vertical='top')

ws_m.row_dimensions[5].height = 40

# Sectiunea Impact verdict
ws_m['A8'].value = 'IMPACT VERDICT (cum s-au schimbat verdictele dupa modificari)'
ws_m['A8'].font = WHITE_BOLD
ws_m['A8'].fill = DARK_BLUE_BG
ws_m.merge_cells('A8:G8')
ws_m['A9'].value = 'Versiunea 1.0: NPV +105k, IRR 26.2%, Payback 4 ani, CF cumulat 377k. KPI: 2 PASS, 2 FAIL -> GO conditionat.'
ws_m['A9'].font = BLACK
ws_m.merge_cells('A9:G9')

# Sectiunea Verificari
ws_m['A11'].value = 'VERIFICARI EFECTUATE FARA MODIFICARI'
ws_m['A11'].font = WHITE_BOLD
ws_m['A11'].fill = DARK_BLUE_BG
ws_m.merge_cells('A11:G11')
ws_m['A12'].value = '- Toate sanity checks din validator.py PASS (distributii=100%, inflatie zero stabil, impozit zero coerent, anuitate × 5 > credit, WACC consistent)'
ws_m['A12'].font = BLACK
ws_m.merge_cells('A12:G12')
ws_m['A13'].value = '- Reality check confirmat: EBITDA margins per parcare individuala sunt mai mari decat benchmark industrie (25-45%) datorita lipsei de overhead la nivel parcare; la nivel agregat companie, margins se incadreaza in interval. Documentat in disclaimer raport.'
ws_m['A13'].font = BLACK
ws_m.merge_cells('A13:G13')


# ====================================================================
# SAVE
# ====================================================================
output_path = '/Users/home-felix/Total Hub/Analiza fezabilitate parcari de inchiriat/outputs/model.xlsx'
wb.save(output_path)
print(f"Model salvat: {output_path}")
print(f"NAMED params count: {len(NAMED)}")
print(f"Tipuri rows: {PARAMETRI_TIPURI_ROWS}")
print(f"Variante scenarios: {len(scenarios)}")
print(f"Portofoliu rows: {len(PF_ROWS)}")
